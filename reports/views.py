"""
Reports Views
"""

from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db.models import Sum, Count, Q, Avg
from django.utils import timezone
from datetime import datetime, timedelta, date
import calendar
from decimal import Decimal
import json

# Import correct financial calculation helpers
from core.financial_helpers import (
    get_branch_balance, get_expected_mission_due, get_mission_income,
    get_mission_obligations, get_local_only_contributions, get_mission_shared_contributions,
    get_financial_summary
)


@login_required
def reports_index(request):
    """Reports dashboard."""
    from core.models import Area, District, Branch, FiscalYear
    from contributions.models import Contribution
    from expenditure.models import Expenditure
    from attendance.models import AttendanceSession
    
    if not (request.user.is_any_admin or request.user.is_auditor):
        messages.error(request, 'Access denied.')
        return redirect('core:dashboard')
    
    # DEPRECATED: Year-as-state architecture - Use current year date range instead of fiscal year
    current_year = timezone.now().year
    from_date = datetime(current_year, 1, 1).date()
    to_date = datetime(current_year, 12, 31).date()
    
    # Apply user hierarchy filtering for stats
    user_scope = 'mission'  # Default for mission admin
    branch_filter = {}
    if request.user.is_branch_executive and request.user.branch:
        branch_filter = {'branch': request.user.branch}
        user_scope = 'branch'
    elif request.user.is_district_executive and request.user.managed_district:
        branch_filter = {'branch__district': request.user.managed_district}
        user_scope = 'district'
    elif request.user.is_area_executive and request.user.managed_area:
        branch_filter = {'branch__district__area': request.user.managed_area}
        user_scope = 'area'
    
    # Quick stats - use date filtering instead of fiscal year filtering
    total_contributions = Contribution.objects.filter(
        date__gte=from_date,
        date__lte=to_date,
        **branch_filter
    ).aggregate(total=Sum('amount'))['total'] or Decimal('0.00')
    
    total_expenditure = Expenditure.objects.filter(
        date__gte=from_date,
        date__lte=to_date,
        **branch_filter if user_scope != 'area' and user_scope != 'district' else {}  # Mission-level expenses for area/district execs
    ).aggregate(total=Sum('amount'))['total'] or Decimal('0.00')
    
    total_branches = Branch.objects.filter(is_active=True).count() if user_scope == 'mission' else (
        Branch.objects.filter(district__area=request.user.managed_area, is_active=True).count() if user_scope == 'area' else
        Branch.objects.filter(district=request.user.managed_district, is_active=True).count() if user_scope == 'district' else
        1 if user_scope == 'branch' else 0
    )
    total_sessions = AttendanceSession.objects.filter(
        date__year=current_year,
        **branch_filter
    ).count()
    
    context = {
        'areas': Area.objects.filter(pk=request.user.managed_area.pk, is_active=True) if user_scope == 'area' else Area.objects.filter(is_active=True),
        'districts': District.objects.filter(area=request.user.managed_area, is_active=True) if user_scope == 'area' else (
            District.objects.filter(pk=request.user.managed_district.pk, is_active=True) if user_scope == 'district' else (
                District.objects.none() if user_scope == 'branch' else District.objects.filter(is_active=True)
            )
        ),
        'branches': Branch.objects.filter(district__area=request.user.managed_area, is_active=True) if user_scope == 'area' else (
            Branch.objects.filter(district=request.user.managed_district, is_active=True) if user_scope == 'district' else (
                Branch.objects.filter(pk=request.user.branch.pk, is_active=True) if user_scope == 'branch' else Branch.objects.filter(is_active=True)
            )
        ),
        'current_year': current_year,  # Use current year instead of fiscal year object
        'total_contributions': total_contributions,
        'total_expenditure': total_expenditure,
        'total_branches': total_branches,
        'total_sessions': total_sessions,
    }
    return render(request, 'reports/index.html', context)


@login_required
def contribution_report(request):
    """Contribution reports with filtering and charts."""
    from core.models import Area, District, Branch, FiscalYear, SiteSettings
    from contributions.models import Contribution, ContributionType
    
    if not (request.user.is_any_admin or request.user.is_auditor):
        messages.error(request, 'Access denied.')
        return redirect('core:dashboard')
    
    # Get filters
    area_id = request.GET.get('area')
    district_id = request.GET.get('district')
    branch_id = request.GET.get('branch')
    from_date = request.GET.get('from_date')
    to_date = request.GET.get('to_date')
    
    # DEPRECATED: Year-as-state architecture - Use date filtering instead of fiscal year
    # Set default date range to current year if no dates provided
    if not from_date or not to_date:
        current_year = timezone.now().year
        from_date = datetime(current_year, 1, 1).date()
        to_date = datetime(current_year, 12, 31).date()
    else:
        from_date = datetime.strptime(from_date, '%Y-%m-%d').date()
        to_date = datetime.strptime(to_date, '%Y-%m-%d').date()
    
    # Base queryset - use date filtering instead of fiscal year filtering
    contributions = Contribution.objects.filter(
        date__gte=from_date,
        date__lte=to_date
    )
    
    # Apply user hierarchy filtering
    user_scope = 'mission'  # Default for mission admin
    if request.user.is_branch_executive and request.user.branch:
        contributions = contributions.filter(branch=request.user.branch)
        user_scope = 'branch'
    elif request.user.is_district_executive and request.user.managed_district:
        contributions = contributions.filter(branch__district=request.user.managed_district)
        user_scope = 'district'
    elif request.user.is_area_executive and request.user.managed_area:
        contributions = contributions.filter(branch__district__area=request.user.managed_area)
        user_scope = 'area'
    
    # Apply filters - respect user hierarchy
    if user_scope == 'area':
        districts = District.objects.filter(area=request.user.managed_area, is_active=True)
        branches = Branch.objects.filter(district__area=request.user.managed_area, is_active=True)
    elif user_scope == 'district':
        districts = District.objects.filter(pk=request.user.managed_district.pk, is_active=True)
        branches = Branch.objects.filter(district=request.user.managed_district, is_active=True)
    elif user_scope == 'branch':
        districts = District.objects.none()
        branches = Branch.objects.filter(pk=request.user.branch.pk, is_active=True)
    else:
        districts = District.objects.filter(is_active=True)
        branches = Branch.objects.filter(is_active=True)
    
    if area_id:
        contributions = contributions.filter(branch__district__area_id=area_id)
        districts = districts.filter(area_id=area_id)
        branches = branches.filter(district__area_id=area_id)
    
    if district_id:
        contributions = contributions.filter(branch__district_id=district_id)
        branches = branches.filter(district_id=district_id)
    
    if branch_id:
        contributions = contributions.filter(branch_id=branch_id)
    
    # Calculate statistics
    total_amount = contributions.aggregate(total=Sum('amount'))['total'] or Decimal('0.00')
    total_count = contributions.count()
    avg_contribution = contributions.aggregate(avg=Avg('amount'))['avg'] or Decimal('0.00')
    
    # By contribution type
    by_type = contributions.values('contribution_type__name').annotate(
        total=Sum('amount'),
        count=Count('id')
    ).order_by('-total')
    
    # By branch
    by_branch = contributions.values('branch__name').annotate(
        total=Sum('amount'),
        count=Count('id')
    ).order_by('-total')[:10]
    
    # Monthly trend
    from django.db.models.functions import Extract
    monthly_data = contributions.annotate(
        month=Extract('date', 'month')
    ).values('month').annotate(
        total=Sum('amount')
    ).order_by('month')
    
    context = {
        'areas': Area.objects.filter(pk=request.user.managed_area.pk, is_active=True) if user_scope == 'area' else Area.objects.filter(is_active=True),
        'districts': districts,
        'branches': branches,
        'contribution_types': ContributionType.objects.filter(is_active=True),
        'total_amount': total_amount,
        'total_count': total_count,
        'avg_contribution': avg_contribution,
        'by_type': by_type,
        'by_branch': by_branch,
        'monthly_data': list(monthly_data),
        # DEPRECATED: Year-as-state architecture - fiscal_year no longer used
        'site_settings': SiteSettings.get_settings(),
    }
    
    return render(request, 'reports/contribution_report.html', context)


@login_required
def expenditure_report(request):
    """Expenditure reports with filtering and charts."""
    from core.models import Area, District, Branch, FiscalYear
    from expenditure.models import Expenditure, ExpenditureCategory
    
    if not (request.user.is_any_admin or request.user.is_auditor):
        messages.error(request, 'Access denied.')
        return redirect('core:dashboard')
    
    # Get filters
    area_id = request.GET.get('area')
    district_id = request.GET.get('district')
    branch_id = request.GET.get('branch')
    from_date = request.GET.get('from_date')
    to_date = request.GET.get('to_date')
    
    # DEPRECATED: Year-as-state architecture - Use date filtering instead of fiscal year
    # Set default date range to current year if no dates provided
    if not from_date or not to_date:
        current_year = timezone.now().year
        from_date = datetime(current_year, 1, 1).date()
        to_date = datetime(current_year, 12, 31).date()
    else:
        from_date = datetime.strptime(from_date, '%Y-%m-%d').date()
        to_date = datetime.strptime(to_date, '%Y-%m-%d').date()
    
    # Base queryset - use date filtering instead of fiscal year filtering
    expenditures = Expenditure.objects.filter(
        date__gte=from_date,
        date__lte=to_date
    )
    
    # Apply user hierarchy filtering
    user_scope = 'mission'  # Default for mission admin
    if request.user.is_branch_executive and request.user.branch:
        expenditures = expenditures.filter(branch=request.user.branch)
        user_scope = 'branch'
    elif request.user.is_district_executive and request.user.managed_district:
        expenditures = expenditures.filter(Q(branch__district=request.user.managed_district) | Q(level='mission'))
        user_scope = 'district'
    elif request.user.is_area_executive and request.user.managed_area:
        expenditures = expenditures.filter(Q(branch__district__area=request.user.managed_area) | Q(level='mission'))
        user_scope = 'area'
    
    # Apply filters - respect user hierarchy
    if user_scope == 'area':
        districts = District.objects.filter(area=request.user.managed_area, is_active=True)
        branches = Branch.objects.filter(district__area=request.user.managed_area, is_active=True)
    elif user_scope == 'district':
        districts = District.objects.filter(pk=request.user.managed_district.pk, is_active=True)
        branches = Branch.objects.filter(district=request.user.managed_district, is_active=True)
    elif user_scope == 'branch':
        districts = District.objects.none()
        branches = Branch.objects.filter(pk=request.user.branch.pk, is_active=True)
    else:
        districts = District.objects.filter(is_active=True)
        branches = Branch.objects.filter(is_active=True)
    
    if area_id:
        expenditures = expenditures.filter(Q(branch__district__area_id=area_id) | Q(level='mission'))
        districts = districts.filter(area_id=area_id)
        branches = branches.filter(district__area_id=area_id)
    
    if district_id:
        expenditures = expenditures.filter(Q(branch__district_id=district_id) | Q(level='mission'))
        branches = branches.filter(district_id=district_id)
    
    if branch_id:
        expenditures = expenditures.filter(branch_id=branch_id)
    
    if from_date:
        expenditures = expenditures.filter(date__gte=from_date)
    if to_date:
        expenditures = expenditures.filter(date__lte=to_date)
    
    # Calculate statistics
    total_amount = expenditures.aggregate(total=Sum('amount'))['total'] or Decimal('0.00')
    total_count = expenditures.count()
    approved_amount = expenditures.filter(
        Q(status='approved') | Q(status='paid')
    ).aggregate(total=Sum('amount'))['total'] or Decimal('0.00')
    pending_amount = expenditures.filter(status='pending').aggregate(total=Sum('amount'))['total'] or Decimal('0.00')
    
    # By category
    by_category = expenditures.values('category__name').annotate(
        total=Sum('amount'),
        count=Count('id')
    ).order_by('-total')
    
    # By level
    by_level = expenditures.values('level').annotate(
        total=Sum('amount'),
        count=Count('id')
    ).order_by('-total')
    
    # By branch
    by_branch = expenditures.filter(branch__isnull=False).values('branch__name').annotate(
        total=Sum('amount'),
        count=Count('id')
    ).order_by('-total')[:10]
    
    context = {
        'areas': Area.objects.filter(pk=request.user.managed_area.pk, is_active=True) if user_scope == 'area' else Area.objects.filter(is_active=True),
        'districts': districts,
        'branches': branches,
        'categories': ExpenditureCategory.objects.filter(is_active=True),
        'total_amount': total_amount,
        'total_count': total_count,
        'approved_amount': approved_amount,
        'pending_amount': pending_amount,
        'by_category': by_category,
        'by_level': by_level,
        'by_branch': by_branch,
        # DEPRECATED: Year-as-state architecture - fiscal_year no longer used
    }
    
    return render(request, 'reports/expenditure_report.html', context)


@login_required
def attendance_report(request):
    """Attendance reports with filtering and statistics."""
    from core.models import Area, District, Branch
    from attendance.models import AttendanceSession, AttendanceRecord, VisitorRecord
    from accounts.models import User
    from django.db.models.functions import Extract
    import json
    
    if not (request.user.is_any_admin or request.user.is_auditor):
        messages.error(request, 'Access denied.')
        return redirect('core:dashboard')
    
    # Get filters
    area_id = request.GET.get('area')
    district_id = request.GET.get('district')
    branch_id = request.GET.get('branch')
    from_date = request.GET.get('start_date')
    to_date = request.GET.get('end_date')
    
    # Base queryset
    sessions = AttendanceSession.objects.all().select_related('branch', 'branch__district', 'branch__district__area')
    
    # Get all options for dropdowns (with relationships for JS filtering)
    all_areas = Area.objects.filter(is_active=True)
    all_districts = District.objects.filter(is_active=True).select_related('area')
    all_branches = Branch.objects.filter(is_active=True).select_related('district')
    
    # Apply filters to sessions
    filtered_districts = all_districts
    filtered_branches = all_branches
    
    if area_id:
        sessions = sessions.filter(branch__district__area_id=area_id)
        filtered_districts = all_districts.filter(area_id=area_id)
        filtered_branches = all_branches.filter(district__area_id=area_id)
    
    if district_id:
        sessions = sessions.filter(branch__district_id=district_id)
        filtered_branches = filtered_branches.filter(district_id=district_id)
    
    if branch_id:
        sessions = sessions.filter(branch_id=branch_id)
    
    if from_date:
        sessions = sessions.filter(date__gte=from_date)
    if to_date:
        sessions = sessions.filter(date__lte=to_date)
    
    # Calculate statistics
    total_sessions = sessions.count()
    total_attendance = AttendanceRecord.objects.filter(
        session__in=sessions,
        status='present'
    ).count()
    
    # Visitors count
    total_visitors = VisitorRecord.objects.filter(
        session__in=sessions
    ).count() if total_sessions > 0 else 0
    
    # Peak attendance
    peak_session = sessions.order_by('-total_attendance').first()
    peak_attendance = peak_session.total_attendance if peak_session else 0
    
    # By branch
    by_branch = sessions.values('branch__name').annotate(
        session_count=Count('id'),
        total_present=Count('attendance_records', filter=Q(attendance_records__status='present'))
    ).order_by('-session_count')[:10]
    
    # Average attendance per session
    avg_attendance = total_attendance / total_sessions if total_sessions > 0 else 0
    
    # Total members
    total_members = User.objects.filter(role='member', is_active=True).count()
    
    # Chart data - weekly attendance (last 4 weeks or by date range)
    weekly_data = sessions.annotate(
        week=Extract('date', 'week')
    ).values('week').annotate(
        total=Count('attendance_records', filter=Q(attendance_records__status='present'))
    ).order_by('week')[:8]
    
    chart_labels = [f"Week {d['week']}" for d in weekly_data] if weekly_data else ['Week 1', 'Week 2', 'Week 3', 'Week 4']
    chart_data = [d['total'] for d in weekly_data] if weekly_data else [0, 0, 0, 0]
    
    context = {
        'areas': all_areas,
        'districts': all_districts,
        'branches': all_branches,
        'selected_area': area_id,
        'selected_district': district_id,
        'selected_branch': branch_id,
        'total_sessions': total_sessions,
        'total_attendance': total_attendance,
        'avg_attendance': avg_attendance,
        'total_members': total_members,
        'total_visitors': total_visitors,
        'peak_attendance': peak_attendance,
        'by_branch': by_branch,
        'chart_labels': json.dumps(chart_labels),
        'chart_data': json.dumps(chart_data),
    }
    
    return render(request, 'reports/attendance_report.html', context)


@login_required
def financial_report(request):
    """Comprehensive financial summary reports."""
    from core.models import Area, District, Branch, FiscalYear
    from contributions.models import Contribution, ContributionType
    from expenditure.models import Expenditure, ExpenditureCategory
    from .models_hierarchy import AreaFinancialReport, DistrictFinancialReport
    
    if not (request.user.is_any_admin or request.user.is_auditor):
        messages.error(request, 'Access denied.')
        return redirect('core:dashboard')
    
    # Get filters
    area_id = request.GET.get('area')
    district_id = request.GET.get('district')
    branch_id = request.GET.get('branch')
    from_date = request.GET.get('from_date')
    to_date = request.GET.get('to_date')
    
    # DEPRECATED: Year-as-state architecture - Use date filtering instead of fiscal year
    # Set default date range to current year if no dates provided
    if not from_date or not to_date:
        current_year = timezone.now().year
        from_date = datetime(current_year, 1, 1).date()
        to_date = datetime(current_year, 12, 31).date()
    else:
        from_date = datetime.strptime(from_date, '%Y-%m-%d').date()
        to_date = datetime.strptime(to_date, '%Y-%m-%d').date()
    
    # Base querysets - use date filtering instead of fiscal year filtering
    contributions = Contribution.objects.filter(
        date__gte=from_date,
        date__lte=to_date
    )
    expenditures = Expenditure.objects.filter(
        date__gte=from_date,
        date__lte=to_date
    )
    
    # Apply filters
    districts = District.objects.filter(is_active=True)
    branches = Branch.objects.filter(is_active=True)
    
    if area_id:
        contributions = contributions.filter(branch__district__area_id=area_id)
        expenditures = expenditures.filter(Q(branch__district__area_id=area_id) | Q(level='mission'))
        districts = districts.filter(area_id=area_id)
        branches = branches.filter(district__area_id=area_id)
    
    if district_id:
        contributions = contributions.filter(branch__district_id=district_id)
        expenditures = expenditures.filter(Q(branch__district_id=district_id) | Q(level='mission'))
        branches = branches.filter(district_id=district_id)
    
    if branch_id:
        contributions = contributions.filter(branch_id=branch_id)
        expenditures = expenditures.filter(branch_id=branch_id)
    
    if from_date:
        contributions = contributions.filter(date__gte=from_date)
        expenditures = expenditures.filter(date__gte=from_date)
    if to_date:
        contributions = contributions.filter(date__lte=to_date)
        expenditures = expenditures.filter(date__lte=to_date)
    
    # Calculate totals
    total_income = contributions.aggregate(total=Sum('amount'))['total'] or Decimal('0.00')
    total_expense = expenditures.aggregate(total=Sum('amount'))['total'] or Decimal('0.00')
    net_balance = total_income - total_expense
    
    # By contribution type
    income_by_type = contributions.values('contribution_type__name').annotate(
        total=Sum('amount')
    ).order_by('-total')
    
    # By expenditure category
    expense_by_category = expenditures.values('category__name').annotate(
        total=Sum('amount')
    ).order_by('-total')
    
    # Prepare chart data
    income_chart_data = {
        'labels': [item['contribution_type__name'] for item in income_by_type],
        'data': [float(item['total']) for item in income_by_type]
    }
    
    expense_chart_data = {
        'labels': [item['category__name'] for item in expense_by_category],
        'data': [float(item['total']) for item in expense_by_category]
    }
    
    context = {
        'areas': Area.objects.filter(is_active=True),
        'districts': districts,
        'branches': branches,
        # DEPRECATED: Year-as-state architecture - fiscal_year no longer used
        'total_income': total_income,
        'total_expense': total_expense,
        'net_balance': net_balance,
        'income_by_type': income_by_type,
        'expense_by_category': expense_by_category,
        'income_chart_data': json.dumps(income_chart_data),
        'expense_chart_data': json.dumps(expense_chart_data),
        # Add generated reports
        'area_reports': AreaFinancialReport.objects.all().order_by('-year', '-month')[:10] if request.user.is_mission_admin or request.user.is_auditor or request.user.is_area_executive else AreaFinancialReport.objects.none(),
        'district_reports': DistrictFinancialReport.objects.all().order_by('-year', '-month')[:10] if request.user.is_mission_admin or request.user.is_auditor or request.user.is_area_executive or request.user.is_district_executive else DistrictFinancialReport.objects.none(),
    }
    
    return render(request, 'reports/financial_report.html', context)


@login_required
def financial_report_print(request):
    """Print-optimized financial report."""
    from core.models import Area, District, Branch, FiscalYear, SiteSettings
    from contributions.models import Contribution, ContributionType
    from expenditure.models import Expenditure, ExpenditureCategory
    from members.models import Member
    from attendance.models import AttendanceSession, AttendanceRecord
    
    if not (request.user.is_any_admin or request.user.is_auditor):
        messages.error(request, 'Access denied.')
        return redirect('core:dashboard')
    
    # Get filters (same as main report)
    area_id = request.GET.get('area')
    district_id = request.GET.get('district')
    branch_id = request.GET.get('branch')
    from_date = request.GET.get('from_date')
    to_date = request.GET.get('to_date')
    
    # DEPRECATED: Year-as-state architecture - Use date filtering instead of fiscal year
    # Set default date range to current year if no dates provided
    if not from_date or not to_date:
        current_year = timezone.now().year
        from_date = datetime(current_year, 1, 1).date()
        to_date = datetime(current_year, 12, 31).date()
    else:
        from_date = datetime.strptime(from_date, '%Y-%m-%d').date() if from_date else None
        to_date = datetime.strptime(to_date, '%Y-%m-%d').date() if to_date else None
    
    # Base querysets - use date filtering instead of fiscal year filtering
    contributions = Contribution.objects.filter(
        date__gte=from_date,
        date__lte=to_date
    )
    expenditures = Expenditure.objects.filter(
        date__gte=from_date,
        date__lte=to_date
    )
    
    # Apply filters
    districts = District.objects.filter(is_active=True)
    branches = Branch.objects.filter(is_active=True)
    
    if area_id:
        contributions = contributions.filter(branch__district__area_id=area_id)
        expenditures = expenditures.filter(Q(branch__district__area_id=area_id) | Q(level='mission'))
        districts = districts.filter(area_id=area_id)
        branches = branches.filter(district__area_id=area_id)
    
    if district_id:
        contributions = contributions.filter(branch__district_id=district_id)
        expenditures = expenditures.filter(Q(branch__district_id=district_id) | Q(level='mission'))
        branches = branches.filter(district_id=district_id)
    
    if branch_id:
        contributions = contributions.filter(branch_id=branch_id)
        expenditures = expenditures.filter(branch_id=branch_id)
    
    if from_date:
        contributions = contributions.filter(date__gte=from_date)
        expenditures = expenditures.filter(date__gte=from_date)
    if to_date:
        contributions = contributions.filter(date__lte=to_date)
        expenditures = expenditures.filter(date__lte=to_date)
    
    # Calculate totals
    total_income = contributions.aggregate(total=Sum('amount'))['total'] or Decimal('0.00')
    total_expenditure = expenditures.aggregate(total=Sum('amount'))['total'] or Decimal('0.00')
    net_amount = total_income - total_expenditure
    
    # By contribution type
    income_by_type = contributions.values('contribution_type__name').annotate(
        total=Sum('amount')
    ).order_by('-total')
    
    # By expenditure category
    expense_by_category = expenditures.values('category__name').annotate(
        total=Sum('amount')
    ).order_by('-total')
    
    # Branch performance data
    branch_performance = []
    for branch in branches:
        branch_income = contributions.filter(branch=branch).aggregate(total=Sum('amount'))['total'] or Decimal('0.00')
        branch_expense = expenditures.filter(branch=branch).aggregate(total=Sum('amount'))['total'] or Decimal('0.00')
        branch_net = branch_income - branch_expense
        member_count = Member.objects.filter(user__branch=branch, status='active').count()
        
        branch_performance.append({
            'name': branch.name,
            'district': branch.district.name if branch.district else 'Mission',
            'total_income': branch_income,
            'total_expenditure': branch_expense,
            'net_amount': branch_net,
            'member_count': member_count
        })
    
    # Calculate attendance rate
    total_sessions = AttendanceSession.objects.filter(
        date__year=timezone.now().year
    ).count()
    total_attendance = AttendanceRecord.objects.filter(
        session__date__year=timezone.now().year,
        status='present'
    ).count()
    total_members = Member.objects.filter(status='active').count()
    
    attendance_rate = 0
    if total_members > 0 and total_sessions > 0:
        attendance_rate = (total_attendance / (total_members * total_sessions)) * 100
    
    # Set date range for report
    # DEPRECATED: Year-as-state architecture - Use date filtering only
    start_date = from_date or timezone.now().date()
    end_date = to_date or timezone.now().date()
    current_date = timezone.now()
    
    context = {
        'site_settings': site_settings,
        # DEPRECATED: Year-as-state architecture - fiscal_year no longer used
        'total_income': total_income,
        'total_expenditure': total_expenditure,
        'net_amount': net_amount,
        'income_by_type': income_by_type,
        'expense_by_category': expense_by_category,
        'branch_performance': branch_performance,
        'total_members': total_members,
        'total_branches': branches.count(),
        'attendance_rate': attendance_rate,
        'start_date': start_date,
        'end_date': end_date,
        'current_date': current_date,
    }
    
    return render(request, 'reports/financial_report_print.html', context)


# ============ MONTHLY REPORTS ============

@login_required
def monthly_reports(request):
    """Monthly reports listing and management."""
    from .models import MonthlyReport
    from core.models import Area, District, Branch, FiscalYear
    
    if not (request.user.is_mission_admin or request.user.is_auditor):
        messages.error(request, 'Access denied.')
        return redirect('core:dashboard')
    
    # Get filters
    selected_month = request.GET.get('month', str(timezone.now().month))
    selected_year = request.GET.get('year', str(timezone.now().year))
    selected_branch = request.GET.get('branch', '')
    selected_status = request.GET.get('status', '')
    
    # Convert to int safely
    try:
        selected_month = int(selected_month) if selected_month else timezone.now().month
    except (ValueError, TypeError):
        selected_month = timezone.now().month
    
    try:
        selected_year = int(selected_year) if selected_year else timezone.now().year
    except (ValueError, TypeError):
        selected_year = timezone.now().year
    
    # Build queryset
    reports = MonthlyReport.objects.select_related(
        'branch', 'branch__district', 'branch__district__area',
        'submitted_by', 'approved_by'
    )
    
    if selected_month:
        reports = reports.filter(month=selected_month)
    if selected_year:
        reports = reports.filter(year=selected_year)
    if selected_branch:
        reports = reports.filter(branch_id=selected_branch)
    if selected_status:
        reports = reports.filter(status=selected_status)
    
    # Calculate summary
    total_due = sum(r.mission_remittance_balance for r in reports if r.mission_remittance_balance > 0)
    total_overdue = sum(r.mission_remittance_balance for r in reports if r.is_overdue)
    
    context = {
        'reports': reports,
        'areas': Area.objects.filter(is_active=True),
        'districts': District.objects.filter(is_active=True),
        'branches': Branch.objects.filter(is_active=True),
        'fiscal_years': FiscalYear.objects.all(),
        'months': [(i, timezone.now().replace(day=1, month=i).strftime('%B')) for i in range(1, 13)],
        'years': range(timezone.now().year - 2, timezone.now().year + 1),
        'status_choices': MonthlyReport.Status.choices,
        'selected_month': selected_month,
        'selected_year': selected_year,
        'selected_branch': selected_branch,
        'selected_status': selected_status,
        'total_due': total_due,
        'total_overdue': total_overdue,
    }
    
    return render(request, 'reports/monthly_reports.html', context)


@login_required
def monthly_report_detail(request, pk):
    """View and manage a specific monthly report."""
    from .models import MonthlyReport
    
    report = get_object_or_404(MonthlyReport, pk=pk)
    
    # Check permissions
    if not (request.user.is_mission_admin or request.user.is_auditor or 
            (request.user.is_branch_executive and request.user.branch == report.branch)):
        messages.error(request, 'Access denied.')
        return redirect('reports:monthly_reports')
    
    # Handle actions
    if request.method == 'POST':
        action = request.POST.get('action')
        
        if action == 'approve' and (request.user.is_mission_admin or request.user.is_auditor):
            report.mark_approved(request.user)
            messages.success(request, f'Report for {report.branch.name} approved successfully.')
        
        elif action == 'mark_paid' and (request.user.is_mission_admin or request.user.is_auditor):
            payment_date = request.POST.get('payment_date')
            payment_reference = request.POST.get('payment_reference', '')
            payment_method = request.POST.get('payment_method', '')
            
            if payment_date:
                report.mark_paid(
                    payment_date=payment_date,
                    reference=payment_reference,
                    method=payment_method
                )
                messages.success(request, f'Payment recorded for {report.branch.name}.')
            else:
                messages.error(request, 'Please provide payment date.')
        
        elif action == 'submit' and request.user.is_branch_executive and request.user.branch == report.branch:
            report.mark_submitted(request.user)
            messages.success(request, 'Report submitted for review.')
        
        return redirect('reports:monthly_report_detail', pk=report.pk)
    
    context = {
        'report': report,
        'detail_items': report.detail_items.all(),
        'can_edit': request.user.is_branch_executive and request.user.branch == report.branch,
        'can_approve': request.user.is_mission_admin or request.user.is_auditor,
    }
    
    return render(request, 'reports/monthly_report_detail.html', context)


@login_required
def monthly_report_generate(request):
    """Generate or create a monthly report."""
    from .models import MonthlyReport
    from contributions.models import Contribution
    from expenditure.models import Expenditure
    from attendance.models import AttendanceSession
    from core.models import FiscalYear
    
    # Only branch executives can generate reports for their branch
    if not request.user.is_branch_executive:
        messages.error(request, 'Only branch executives can generate reports.')
        return redirect('reports:monthly_reports')
    
    branch = request.user.branch
    if not branch:
        messages.error(request, 'No branch assigned to your account.')
        return redirect('reports:monthly_reports')
    
    if request.method == 'POST':
        month = int(request.POST.get('month'))
        year = int(request.POST.get('year'))
        # DEPRECATED: Year-as-state architecture - Create/get fiscal year for compatibility
        fiscal_year, _ = FiscalYear.objects.get_or_create(
            year=year,
            defaults={
                'start_date': date(year, 1, 1),
                'end_date': date(year, 12, 31),
                'is_current': False,  # DEPRECATED: Not used
                'is_closed': False
            }
        )
        
        # Check if report already exists
        if MonthlyReport.objects.filter(branch=branch, month=month, year=year).exists():
            messages.warning(request, f'Report for {branch.name} - {month}/{year} already exists.')
            return redirect('reports:monthly_reports')
        
        # Calculate data from existing records
        month_start = timezone.now().replace(year=year, month=month, day=1)
        if month == 12:
            month_end = timezone.now().replace(year=year+1, month=1, day=1) - timedelta(days=1)
        else:
            month_end = timezone.now().replace(year=year, month=month+1, day=1) - timedelta(days=1)
        
        # Get contributions
        # DEPRECATED: Year-as-state architecture - Use date filtering only
        contributions = Contribution.objects.filter(
            branch=branch,
            date__gte=month_start,
            date__lte=month_end
        )
        
        tithe_total = contributions.filter(contribution_type__category='tithe').aggregate(
            total=Sum('amount')
        )['total'] or Decimal('0.00')
        
        offering_total = contributions.filter(contribution_type__category='offering').aggregate(
            total=Sum('amount')
        )['total'] or Decimal('0.00')
        
        special_total = contributions.filter(contribution_type__category='special').aggregate(
            total=Sum('amount')
        )['total'] or Decimal('0.00')
        
        # Get expenditures
        # DEPRECATED: Year-as-state architecture - Use date filtering only
        expenditures = Expenditure.objects.filter(
            branch=branch,
            date__gte=month_start,
            date__lte=month_end
        )
        
        # Get attendance
        sessions = AttendanceSession.objects.filter(
            branch=branch,
            date__gte=month_start,
            date__lte=month_end
        )
        
        total_services = sessions.count()
        total_attendance = sessions.aggregate(total=Sum('total_attendance'))['total'] or 0
        average_attendance = total_attendance / total_services if total_services > 0 else 0
        
        # Create report
        report = MonthlyReport.objects.create(
            branch=branch,
            fiscal_year=fiscal_year,
            month=month,
            year=year,
            total_services=total_services,
            total_attendance=total_attendance,
            average_attendance=average_attendance,
            total_contributions=contributions.aggregate(total=Sum('amount'))['total'] or Decimal('0.00'),
            tithe_amount=tithe_total,
            offering_amount=offering_total,
            special_contributions=special_total,
            total_expenditure=expenditures.aggregate(total=Sum('amount'))['total'] or Decimal('0.00'),
            created_by=request.user
        )
        
        messages.success(request, f'Report generated successfully for {month}/{year}.')
        return redirect('reports:monthly_report_detail', pk=report.pk)
    
    context = {
        'branch': branch,
        'months': [(i, timezone.now().replace(day=1, month=i).strftime('%B')) for i in range(1, 12)],
        'current_year': timezone.now().year,
    }
    
    return render(request, 'reports/monthly_report_generate.html', context)


@login_required
def monthly_report_export_pdf(request, pk=None):
    """Export monthly report as PDF."""
    from .models import MonthlyReport
    from django.http import HttpResponse
    from django.template.loader import render_to_string
    import weasyprint
    
    if pk:
        # Single report
        report = get_object_or_404(MonthlyReport, pk=pk)
        reports = [report]
        filename = f"monthly_report_{report.branch.name}_{report.month}_{report.year}.pdf"
    else:
        # All reports for selected month/year
        month = request.GET.get('month')
        year = request.GET.get('year')
        if not month or not year:
            messages.error(request, 'Please specify month and year.')
            return redirect('reports:monthly_reports')
        
        reports = MonthlyReport.objects.filter(month=month, year=year).select_related('branch')
        filename = f"monthly_reports_{month}_{year}.pdf"
    
    # Generate PDF
    html_string = render_to_string('reports/monthly_report_pdf.html', {'reports': reports})
    html = weasyprint.HTML(string=html_string)
    pdf = html.write_pdf()
    
    response = HttpResponse(pdf, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


@login_required
def comprehensive_statistics(request):
    """Comprehensive statistics dashboard with hierarchical views."""
    from core.models import Area, District, Branch, FiscalYear, SiteSettings
    from contributions.models import Contribution, ContributionType
    from expenditure.models import Expenditure, ExpenditureCategory
    from members.models import Member
    
    if not (request.user.is_any_admin or request.user.is_auditor):
        messages.error(request, 'Access denied.')
        return redirect('core:dashboard')
    
    # Get filters
    area_id = request.GET.get('area')
    district_id = request.GET.get('district')
    branch_id = request.GET.get('branch')
    year = request.GET.get('year', str(timezone.now().year))
    month = request.GET.get('month')
    
    # DEPRECATED: Year-as-state architecture - Use date filtering instead of fiscal year
    # Set default date range to current year if no dates provided
    if not month or not year:
        current_year = timezone.now().year
        from_date = datetime(current_year, 1, 1).date()
        to_date = datetime(current_year, 12, 31).date()
    else:
        from_date = datetime(int(year), 1, 1).date()
        to_date = datetime(int(year), 12, 31).date()
        if month:
            from_date = datetime(int(year), int(month), 1).date()
            to_date = datetime(int(year), int(month), calendar.monthrange(int(year), int(month))[1]).date()
    
    # Base querysets - use date filtering instead of fiscal year filtering
    contributions = Contribution.objects.filter(
        date__gte=from_date,
        date__lte=to_date
    )
    expenditures = Expenditure.objects.filter(
        date__gte=from_date,
        date__lte=to_date
    )
    
    # Apply filters
    if area_id:
        contributions = contributions.filter(branch__district__area_id=area_id)
        expenditures = expenditures.filter(branch__district__area_id=area_id)
        districts = District.objects.filter(area_id=area_id)
        branches = Branch.objects.filter(district__area_id=area_id)
    else:
        districts = District.objects.filter(is_active=True)
        branches = Branch.objects.filter(is_active=True)
    
    if district_id:
        contributions = contributions.filter(branch__district_id=district_id)
        expenditures = expenditures.filter(branch__district_id=district_id)
        branches = branches.filter(district_id=district_id)
    
    if branch_id:
        contributions = contributions.filter(branch_id=branch_id)
        expenditures = expenditures.filter(branch_id=branch_id)
    
    if year:
        contributions = contributions.filter(date__year=year)
        expenditures = expenditures.filter(date__year=year)
    
    if month:
        contributions = contributions.filter(date__month=month)
        expenditures = expenditures.filter(date__month=month)
    
    # Calculate overall statistics using CORRECT financial logic
    total_contributions = contributions.aggregate(total=Sum('amount'))['total'] or Decimal('0.00')
    total_expenditures = expenditures.aggregate(total=Sum('amount'))['total'] or Decimal('0.00')
    net_balance = total_contributions - total_expenditures
    
    # CORRECT: Mission income comes ONLY from verified remittances, not contributions
    mission_income_data = get_mission_income(from_date, to_date)
    actual_mission_income = mission_income_data['total_income']
    
    # CORRECT: Mission obligations are calculated from contribution allocations, not just tithe
    mission_obligations = get_mission_obligations(from_date, to_date)
    
    # CORRECT: Local-only contributions are tracked separately
    local_contributions = get_local_only_contributions(start_date=from_date, end_date=to_date)
    mission_shared_contributions = get_mission_shared_contributions(start_date=from_date, end_date=to_date)
    
    # Calculate total kept at branches (actual cash, not expected)
    # This is total contributions - actual remittances (not expected)
    total_kept_at_branches = total_contributions - actual_mission_income
    
    # Branch-level statistics using CORRECT financial logic
    branch_stats = []
    for branch in branches:
        # Use helper functions for correct calculations
        branch_balance_data = get_branch_balance(branch, from_date, to_date)
        expected_due = get_expected_mission_due(branch, from_date, to_date)
        
        # Get contribution breakdown
        branch_contributions = contributions.filter(branch=branch)
        branch_expenditures = expenditures.filter(branch=branch)
        
        branch_total_contrib = branch_contributions.aggregate(total=Sum('amount'))['total'] or Decimal('0.00')
        branch_total_exp = branch_expenditures.aggregate(total=Sum('amount'))['total'] or Decimal('0.00')
        
        # CORRECT: Branch balance excludes expected mission amounts, includes actual remittances only
        branch_balance = branch_balance_data['branch_balance']
        branch_actual_remitted = branch_balance_data['total_remitted']
        
        # Calculate amount kept at branch (actual cash position)
        branch_kept_at_branch = branch_balance  # This is the actual cash remaining
        
        member_count = Member.objects.filter(user__branch=branch, user__is_active=True).count()
        
        branch_stats.append({
            'branch': branch,
            'total_contributions': branch_total_contrib,
            'total_expenditures': branch_total_exp,
            'expected_mission_due': expected_due,  # Obligation, not cash
            'actual_mission_remitted': branch_actual_remitted,  # Actual cash sent
            'branch_kept_at_branch': branch_kept_at_branch,
            'branch_balance': branch_balance,  # CORRECT: Actual cash position
            'member_count': member_count,
            'avg_contribution_per_member': branch_total_contrib / member_count if member_count > 0 else Decimal('0.00'),
            'financial_health': {
                'remittance_rate': (branch_actual_remitted / expected_due * 100) if expected_due > 0 else Decimal('0.00'),
                'cash_position': branch_balance
            }
        })
    
    # District-level statistics (if not filtered to specific branch)
    district_stats = []
    if not branch_id:
        for district in districts:
            district_branches = branches.filter(district=district)
            district_contributions = contributions.filter(branch__district=district)
            district_expenditures = expenditures.filter(branch__district=district)
            
            district_total_contrib = district_contributions.aggregate(total=Sum('amount'))['total'] or Decimal('0.00')
            district_total_exp = district_expenditures.aggregate(total=Sum('amount'))['total'] or Decimal('0.00')
            district_tithe = district_contributions.filter(contribution_type__category='tithe').aggregate(total=Sum('amount'))['total'] or Decimal('0.00')
            district_remittance = district_tithe * Decimal('0.10')
            
            district_stats.append({
                'district': district,
                'total_contributions': district_total_contrib,
                'total_expenditures': district_total_exp,
                'tithe_amount': district_tithe,
                'mission_remittance': district_remittance,
                'branch_count': district_branches.count(),
                'total_members': Member.objects.filter(user__branch__district=district, user__is_active=True).count()
            })
    
    # Area-level statistics (if not filtered to specific district)
    area_stats = []
    if not district_id and not branch_id:
        areas = Area.objects.filter(is_active=True)
        for area in areas:
            area_districts = districts.filter(area=area)
            area_contributions = contributions.filter(branch__district__area=area)
            area_expenditures = expenditures.filter(branch__district__area=area)
            
            area_total_contrib = area_contributions.aggregate(total=Sum('amount'))['total'] or Decimal('0.00')
            area_total_exp = area_expenditures.aggregate(total=Sum('amount'))['total'] or Decimal('0.00')
            area_tithe = area_contributions.filter(contribution_type__category='tithe').aggregate(total=Sum('amount'))['total'] or Decimal('0.00')
            area_remittance = area_tithe * Decimal('0.10')
            
            area_stats.append({
                'area': area,
                'total_contributions': area_total_contrib,
                'total_expenditures': area_total_exp,
                'tithe_amount': area_tithe,
                'mission_remittance': area_remittance,
                'district_count': area_districts.count(),
                'total_members': Member.objects.filter(user__branch__district__area=area, user__is_active=True).count()
            })
    
    context = {
        'areas': Area.objects.filter(is_active=True),
        'districts': districts,
        'branches': branches,
        'selected_area': area_id,
        'selected_district': district_id,
        'selected_branch': branch_id,
        'selected_year': year,
        'selected_month': month,
        'years': range(timezone.now().year - 3, timezone.now().year + 2),  # Last 3 years + next year
        
        # CORRECT: Overall stats using proper financial logic
        'total_contributions': total_contributions,
        'total_expenditures': total_expenditures,
        'net_balance': net_balance,
        
        # CORRECT: Mission financials - separate obligations from actual income
        'actual_mission_income': actual_mission_income,  # From verified remittances only
        'total_mission_obligations': mission_obligations,  # Expected amounts from contributions
        'mission_income_gap': mission_obligations - actual_mission_income,  # Underpaid amount
        'remittance_count': mission_income_data['remittance_count'],
        
        # CORRECT: Contribution breakdowns
        'local_only_contributions': local_contributions['total_amount'],
        'mission_shared_contributions': mission_shared_contributions['total_amount'],
        'total_kept_at_branches': total_kept_at_branches,  # Actual cash at branches
        
        # Hierarchical stats
        'branch_stats': branch_stats,
        'district_stats': district_stats,
        'area_stats': area_stats,
        
        # DEPRECATED: Year-as-state architecture - fiscal_year no longer used
        'site_settings': SiteSettings.get_settings(),
    }
    
    return render(request, 'reports/comprehensive_statistics.html', context)


@login_required
def member_contributions_report(request):
    """Individual member contributions report by branch."""
    from core.models import Area, District, Branch, FiscalYear, SiteSettings
    from contributions.models import Contribution, ContributionType
    from members.models import Member
    
    if not (request.user.is_any_admin or request.user.is_auditor):
        messages.error(request, 'Access denied.')
        return redirect('core:dashboard')
    
    # Get filters
    area_id = request.GET.get('area')
    district_id = request.GET.get('district')
    branch_id = request.GET.get('branch')
    from_date = request.GET.get('from_date')
    to_date = request.GET.get('to_date')
    
    # DEPRECATED: Year-as-state architecture - Use date filtering instead of fiscal year
    # Set default date range to current year if no dates provided
    if not from_date or not to_date:
        current_year = timezone.now().year
        from_date = datetime(current_year, 1, 1).date()
        to_date = datetime(current_year, 12, 31).date()
    else:
        from_date = datetime.strptime(from_date, '%Y-%m-%d').date() if from_date else None
        to_date = datetime.strptime(to_date, '%Y-%m-%d').date() if to_date else None
    
    # Base queryset - use date filtering instead of fiscal year filtering
    contributions = Contribution.objects.filter(
        contribution_type__is_individual=True,
        date__gte=from_date,
        date__lte=to_date
    )
    
    # Apply filters
    districts = District.objects.filter(is_active=True)
    branches = Branch.objects.filter(is_active=True)
    
    if area_id:
        contributions = contributions.filter(branch__district__area_id=area_id)
        districts = districts.filter(area_id=area_id)
        branches = branches.filter(district__area_id=area_id)
    
    if district_id:
        contributions = contributions.filter(branch__district_id=district_id)
        branches = branches.filter(district_id=district_id)
    
    if branch_id:
        contributions = contributions.filter(branch_id=branch_id)
    
    if from_date:
        contributions = contributions.filter(date__gte=from_date)
    if to_date:
        contributions = contributions.filter(date__lte=to_date)
    
    # Group by member and branch
    member_stats = contributions.values(
        'member__id', 'member__first_name', 'member__last_name', 'member__member_id',
        'branch__name', 'branch__code'
    ).annotate(
        total_amount=Sum('amount'),
        contribution_count=Count('id'),
        tithe_amount=Sum('amount', filter=Q(contribution_type__category='tithe')),
        offering_amount=Sum('amount', filter=Q(contribution_type__category='offering'))
    ).order_by('-total_amount')
    
    # Calculate totals
    total_amount = contributions.aggregate(total=Sum('amount'))['total'] or Decimal('0.00')
    total_members = member_stats.count()
    
    context = {
        'areas': Area.objects.filter(is_active=True),
        'districts': districts,
        'branches': branches,
        'selected_area': area_id,
        'selected_district': district_id,
        'selected_branch': branch_id,
        'from_date': from_date,
        'to_date': to_date,
        
        'member_stats': member_stats,
        'total_amount': total_amount,
        'total_members': total_members,
        'current_year': timezone.now().year,  # Use current year instead of fiscal year
        'site_settings': SiteSettings.get_settings(),
    }
    
    return render(request, 'reports/member_contributions.html', context)


@login_required
def export_statistics_excel(request):
    """Export comprehensive statistics to Excel."""
    from core.models import Area, District, Branch, FiscalYear
    from contributions.models import Contribution
    from expenditure.models import Expenditure
    from members.models import Member
    
    if not (request.user.is_any_admin or request.user.is_auditor):
        messages.error(request, 'Access denied.')
        return redirect('core:dashboard')
    
    # Get filters
    area_id = request.GET.get('area')
    district_id = request.GET.get('district')
    branch_id = request.GET.get('branch')
    year = request.GET.get('year', str(timezone.now().year))
    month = request.GET.get('month')
    
    # DEPRECATED: Year-as-state architecture - Use date filtering instead of fiscal year
    # Set date range for export
    from_date = date(int(year), 1, 1)
    to_date = date(int(year), 12, 31)
    if month:
        from_date = date(int(year), int(month), 1)
        to_date = date(int(year), int(month), calendar.monthrange(int(year), int(month))[1])
    
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        messages.error(request, 'Excel export requires openpyxl package to be installed.')
        return redirect('reports:comprehensive_statistics')
    
    # Create workbook
    wb = openpyxl.Workbook()
    
    # Remove default sheet
    wb.remove(wb.active)
    
    # Summary sheet
    ws_summary = wb.create_sheet("Summary")
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Summary headers
    headers = ["Metric", "Amount", "Description"]
    for col, header in enumerate(headers, 1):
        cell = ws_summary.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Get data - use date filtering
    contributions = Contribution.objects.filter(date__gte=from_date, date__lte=to_date)
    expenditures = Expenditure.objects.filter(date__gte=from_date, date__lte=to_date)
    
    # Apply filters
    if area_id:
        contributions = contributions.filter(branch__district__area_id=area_id)
        expenditures = expenditures.filter(branch__district__area_id=area_id)
    
    if district_id:
        contributions = contributions.filter(branch__district_id=district_id)
        expenditures = expenditures.filter(branch__district_id=district_id)
    
    if branch_id:
        contributions = contributions.filter(branch_id=branch_id)
        expenditures = expenditures.filter(branch_id=branch_id)
    
    if year:
        contributions = contributions.filter(date__year=year)
        expenditures = expenditures.filter(date__year=year)
    
    if month:
        contributions = contributions.filter(date__month=month)
        expenditures = expenditures.filter(date__month=month)
    
    # Calculate statistics
    total_contributions = contributions.aggregate(total=Sum('amount'))['total'] or Decimal('0.00')
    total_expenditures = expenditures.aggregate(total=Sum('amount'))['total'] or Decimal('0.00')
    total_tithe = contributions.filter(contribution_type__category='tithe').aggregate(total=Sum('amount'))['total'] or Decimal('0.00')
    mission_remittance = total_tithe * Decimal('0.10')
    net_balance = total_contributions - total_expenditures
    
    # Summary data
    summary_data = [
        ["Total Contributions", total_contributions, "All contributions collected"],
        ["Total Expenditures", total_expenditures, "All expenses incurred"],
        ["Total Tithe", total_tithe, "Tithe contributions"],
        ["Mission Remittance Due", mission_remittance, "10% of tithe due to mission"],
        ["Net Balance", net_balance, "Contributions minus expenditures"],
        ["Number of Contributions", contributions.count(), "Total contribution entries"],
        ["Number of Expenditures", expenditures.count(), "Total expenditure entries"]
    ]
    
    for row, data in enumerate(summary_data, 2):
        for col, value in enumerate(data, 1):
            cell = ws_summary.cell(row=row, column=col, value=value)
            if col == 2:  # Amount column
                cell.number_format = '#,##0.00'
    
    # Branch Details Sheet
    ws_branches = wb.create_sheet("Branch Details", 1)
    
    branch_headers = ["Branch Code", "Branch Name", "District", "Area", "Total Contributions", 
                     "Total Expenditures", "Tithe Amount", "Mission Remittance", "Branch Balance", "Member Count"]
    
    for col, header in enumerate(branch_headers, 1):
        cell = ws_branches.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Get branch data
    branches = Branch.objects.filter(is_active=True)
    if area_id:
        branches = branches.filter(district__area_id=area_id)
    if district_id:
        branches = branches.filter(district_id=district_id)
    if branch_id:
        branches = branches.filter(id=branch_id)
    
    for row, branch in enumerate(branches, 2):
        branch_contributions = contributions.filter(branch=branch)
        branch_expenditures = expenditures.filter(branch=branch)
        
        branch_total_contrib = branch_contributions.aggregate(total=Sum('amount'))['total'] or Decimal('0.00')
        branch_total_exp = branch_expenditures.aggregate(total=Sum('amount'))['total'] or Decimal('0.00')
        branch_tithe = branch_contributions.filter(contribution_type__category='tithe').aggregate(total=Sum('amount'))['total'] or Decimal('0.00')
        branch_remittance = branch_tithe * Decimal('0.10')
        branch_balance = branch_total_contrib - branch_total_exp - branch_remittance
        member_count = Member.objects.filter(user__branch=branch, user__is_active=True).count()
        
        branch_data = [
            branch.code,
            branch.name,
            branch.district.name if branch.district else "N/A",
            branch.district.area.name if branch.district and branch.district.area else "N/A",
            branch_total_contrib,
            branch_total_exp,
            branch_tithe,
            branch_remittance,
            branch_balance,
            member_count
        ]
        
        for col, value in enumerate(branch_data, 1):
            cell = ws_branches.cell(row=row, column=col, value=value)
            if col >= 5 and col <= 9:  # Amount columns
                cell.number_format = '#,##0.00'
    
    # Auto-adjust column widths
    for ws in [ws_summary, ws_branches]:
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"statistics_report_{year}_{month or 'all'}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    wb.save(response)
    return response


@login_required
def final_financial_report(request):
    """Comprehensive Income vs Expenditure Final Financial Report."""
    from core.models import Area, District, Branch, FiscalYear, SiteSettings, MissionFinancialSummary, BranchFinancialSummary
    from contributions.models import Contribution, ContributionType, Remittance
    from expenditure.models import Expenditure, ExpenditureCategory
    from payroll.models import PayrollRun, PaySlip
    from decimal import Decimal
    
    if not (request.user.is_mission_admin or request.user.is_auditor):
        messages.error(request, 'Access denied.')
        return redirect('core:dashboard')
    
    # Get filters
    month = request.GET.get('month', str(timezone.now().month))
    year = request.GET.get('year', str(timezone.now().year))
    # DEPRECATED: Year-as-state architecture - Removed fiscal_year usage
    # Monthly reports now use date filtering only
    site_settings = SiteSettings.get_settings()
    
    # Calculate date range
    month_int = int(month)
    year_int = int(year)
    month_start = timezone.now().replace(year=year_int, month=month_int, day=1)
    if month_int == 12:
        month_end = timezone.now().replace(year=year_int+1, month=1, day=1) - timedelta(days=1)
    else:
        month_end = timezone.now().replace(year=year_int, month=month_int+1, day=1) - timedelta(days=1)
    
    # ============ MISSION-LEVEL INCOME ============
    # CORRECT: Mission income comes ONLY from verified remittances, not contributions
    start_date = month_start.date()
    end_date = month_end.date()
    
    mission_income_data = get_mission_income(start_date, end_date)
    total_remittances = mission_income_data['total_income']
    
    # Other mission income (if any) - mission-level contributions only
    mission_other_income = Contribution.objects.filter(
        date__gte=start_date,
        date__lte=end_date,
        branch__isnull=True,  # Mission-level contributions only
    ).aggregate(total=Sum('amount'))['total'] or Decimal('0.00')
    
    total_mission_income = total_remittances + mission_other_income
    
    # ============ MISSION-LEVEL EXPENDITURE ============
    # Payroll
    payroll_runs = PayrollRun.objects.filter(
        month=month_int,
        year=year_int
    )
    total_payroll = Decimal('0.00')
    for run in payroll_runs:
        total_payroll += run.total_net_pay or Decimal('0.00')
    
    # Mission expenses
    mission_expenses = Expenditure.objects.filter(
        level='mission',
        date__gte=month_start,
        date__lte=month_end,
        status__in=['approved', 'paid']
    ).aggregate(total=Sum('amount'))['total'] or Decimal('0.00')
    
    # Mission returns to branches (if any)
    mission_returns = Decimal('0.00')  # Track separately in future
    
    total_mission_expenditure = total_payroll + mission_expenses + mission_returns
    
    # Mission net balance
    mission_net_balance = total_mission_income - total_mission_expenditure
    
    # ============ BRANCH-LEVEL AGGREGATES ============
    branches = Branch.objects.filter(is_active=True)
    
    branch_data = []
    total_branch_income = Decimal('0.00')
    total_branch_tithe = Decimal('0.00')
    total_branch_offerings = Decimal('0.00')
    total_branch_other = Decimal('0.00')
    total_branch_expenses = Decimal('0.00')
    total_branch_remitted = Decimal('0.00')
    total_branch_commission = Decimal('0.00')
    total_branch_actual_remitted = Decimal('0.00')
    total_branch_expenditure = Decimal('0.00')
    
    for branch in branches:
        # Income
        # DEPRECATED: Year-as-state architecture - Use date filtering only
        branch_contributions = Contribution.objects.filter(
            branch=branch,
            date__gte=month_start,
            date__lte=month_end
        )
        
        branch_tithe = branch_contributions.filter(
            contribution_type__category='tithe'
        ).aggregate(total=Sum('amount'))['total'] or Decimal('0.00')
        
        branch_offerings = branch_contributions.filter(
            contribution_type__category='offering'
        ).aggregate(total=Sum('amount'))['total'] or Decimal('0.00')
        
        branch_other_contrib = branch_contributions.exclude(
            contribution_type__category__in=['tithe', 'offering']
        ).aggregate(total=Sum('amount'))['total'] or Decimal('0.00')
        
        branch_total_income = branch_tithe + branch_offerings + branch_other_contrib
        
        # Expenditure
        branch_expenditures = Expenditure.objects.filter(
            branch=branch,
            level='branch',
            date__gte=month_start,
            date__lte=month_end,
            status__in=['approved', 'paid']
        ).aggregate(total=Sum('amount'))['total'] or Decimal('0.00')
        
        # Remittance
        # DEPRECATED: Year-as-state architecture - Use month/year filtering for remittances
        branch_remittance = Remittance.objects.filter(
            branch=branch,
            month=month_int,
            year=year_int,
            status='approved'
        ).aggregate(total=Sum('amount_sent'))['total'] or Decimal('0.00')
        
        # CORRECT: Use actual remittances, not expected amounts
        branch_balance_data = get_branch_balance(branch, start_date, end_date)
        branch_actual_remitted = branch_balance_data['total_remitted']
        
        # Commission (if tracked)
        branch_commission = Decimal('0.00')  # Future: get from commission model
        
        branch_total_exp = branch_expenditures + branch_actual_remitted + branch_commission
        branch_balance = branch_total_income - branch_total_exp
        
        branch_data.append({
            'branch': branch,
            'tithe': branch_tithe,
            'offerings': branch_offerings,
            'other_income': branch_other_contrib,
            'total_income': branch_total_income,
            'expenses': branch_expenditures,
            'remittance': branch_actual_remitted,  # For template compatibility
            'expected_due': get_expected_mission_due(branch, start_date, end_date),  # Obligation
            'commission': branch_commission,
            'total_expenditure': branch_total_exp,
            'balance': branch_balance
        })
        
        # Aggregate totals
        total_branch_income += branch_total_income
        total_branch_tithe += branch_tithe
        total_branch_offerings += branch_offerings
        total_branch_other += branch_other_contrib
        total_branch_expenses += branch_expenditures
        total_branch_actual_remitted += branch_actual_remitted
        total_branch_commission += branch_commission
        total_branch_expenditure += branch_total_exp
    
    total_branch_balance = total_branch_income - total_branch_expenditure
    
    # ============ OVERALL TOTALS ============
    overall_income = total_mission_income + total_branch_income
    overall_expenditure = total_mission_expenditure + total_branch_expenditure
    overall_balance = overall_income - overall_expenditure
    
    # ============ CHARTS DATA ============
    # Mission Income Breakdown
    mission_income_chart = {
        'labels': ['Remittances', 'Other Income'],
        'data': [float(total_remittances), float(mission_other_income)]
    }
    
    # Mission Expenditure Breakdown
    mission_expense_chart = {
        'labels': ['Payroll', 'Mission Expenses', 'Returns to Branches'],
        'data': [float(total_payroll), float(mission_expenses), float(mission_returns)]
    }
    
    # Branch Income Breakdown
    branch_income_chart = {
        'labels': ['Tithe', 'Offerings', 'Other'],
        'data': [float(total_branch_tithe), float(total_branch_offerings), float(total_branch_other)]
    }
    
    # Branch Expenditure Breakdown
    branch_expense_chart = {
        'labels': ['Expenses', 'Remittance', 'Commission'],
        'data': [float(total_branch_expenses), float(total_branch_remitted), float(total_branch_commission)]
    }
    
    # Top 10 branches by income
    top_branches = sorted(branch_data, key=lambda x: x['total_income'], reverse=True)[:10]
    top_branches_chart = {
        'labels': [b['branch'].name for b in top_branches],
        'data': [float(b['total_income']) for b in top_branches]
    }
    
    context = {
        'month': month_int,
        'year': year_int,
        'month_name': month_start.strftime('%B'),
        # DEPRECATED: Year-as-state architecture - fiscal_year no longer used
        'site_settings': site_settings,
        'months': [(i, timezone.now().replace(day=1, month=i).strftime('%B')) for i in range(1, 13)],
        'years': range(timezone.now().year - 2, timezone.now().year + 1),
        
        # Mission Level
        'total_remittances': total_remittances,
        'mission_other_income': mission_other_income,
        'total_mission_income': total_mission_income,
        'total_payroll': total_payroll,
        'mission_expenses': mission_expenses,
        'mission_returns': mission_returns,
        'total_mission_expenditure': total_mission_expenditure,
        'mission_net_balance': mission_net_balance,
        
        # Branch Aggregates
        'total_branch_income': total_branch_income,
        'total_branch_tithe': total_branch_tithe,
        'total_branch_offerings': total_branch_offerings,
        'total_branch_other': total_branch_other,
        'total_branch_expenses': total_branch_expenses,
        'total_branch_remitted': total_branch_actual_remitted,  # Use actual remitted
        'total_branch_commission': total_branch_commission,
        'total_branch_expenditure': total_branch_expenditure,
        'total_branch_balance': total_branch_balance,
        
        # Overall
        'overall_income': overall_income,
        'overall_expenditure': overall_expenditure,
        'overall_balance': overall_balance,
        
        # Branch Details
        'branch_data': branch_data,
        'branch_data_json': json.dumps([
            {
                'branch': {'name': item['branch'].name},
                'tithe': str(item['tithe']),
                'offerings': str(item['offerings']),
                'other_income': str(item['other_income']),
                'total_income': str(item['total_income']),
                'expenses': str(item['expenses']),
                'remittance': str(item['remittance']),
                'commission': str(item['commission']),
                'total_expenditure': str(item['total_expenditure']),
                'balance': str(item['balance'])
            } for item in branch_data
        ]),
        
        # Charts
        'mission_income_chart': json.dumps(mission_income_chart),
        'mission_expense_chart': json.dumps(mission_expense_chart),
        'branch_income_chart': json.dumps(branch_income_chart),
        'branch_expense_chart': json.dumps(branch_expense_chart),
        'top_branches_chart': json.dumps(top_branches_chart),
    }
    
    return render(request, 'reports/final_financial_report.html', context)


@login_required
def export_member_contributions_excel(request):
    """Export member contributions to Excel."""
    from core.models import Area, District, Branch, FiscalYear
    from contributions.models import Contribution
    from members.models import Member
    
    if not (request.user.is_any_admin or request.user.is_auditor):
        messages.error(request, 'Access denied.')
        return redirect('core:dashboard')
    
    # Get filters
    area_id = request.GET.get('area')
    district_id = request.GET.get('district')
    branch_id = request.GET.get('branch')
    from_date = request.GET.get('from_date')
    to_date = request.GET.get('to_date')
    
    # DEPRECATED: Year-as-state architecture - Use date filtering instead of fiscal year
    # Set date range for export
    if not from_date or not to_date:
        current_year = timezone.now().year
        from_date = date(current_year, 1, 1)
        to_date = date(current_year, 12, 31)
    else:
        from_date = datetime.strptime(from_date, '%Y-%m-%d').date()
        to_date = datetime.strptime(to_date, '%Y-%m-%d').date()
    
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        messages.error(request, 'Excel export requires openpyxl package to be installed.')
        return redirect('reports:member_contributions')
    
    # Create workbook
    wb = openpyxl.Workbook()
    
    # Member Contributions Sheet
    ws_members = wb.create_sheet("Member Contributions", 0)
    
    # Headers styling
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Headers
    headers = ["Member ID", "First Name", "Last Name", "Branch Code", "Branch Name", 
              "Total Amount", "Tithe Amount", "Offering Amount", "Contribution Count"]
    
    for col, header in enumerate(headers, 1):
        cell = ws_members.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Get data - use date filtering
    contributions = Contribution.objects.filter(
        date__gte=from_date, 
        date__lte=to_date, 
        contribution_type__is_individual=True
    )
    
    # Apply filters
    if area_id:
        contributions = contributions.filter(branch__district__area_id=area_id)
    if district_id:
        contributions = contributions.filter(branch__district_id=district_id)
    if branch_id:
        contributions = contributions.filter(branch_id=branch_id)
    if from_date:
        contributions = contributions.filter(date__gte=from_date)
    if to_date:
        contributions = contributions.filter(date__lte=to_date)
    
    # Group by member
    member_stats = contributions.values(
        'member__id', 'member__first_name', 'member__last_name', 'member__member_id',
        'branch__name', 'branch__code'
    ).annotate(
        total_amount=Sum('amount'),
        contribution_count=Count('id'),
        tithe_amount=Sum('amount', filter=Q(contribution_type__category='tithe')),
        offering_amount=Sum('amount', filter=Q(contribution_type__category='offering'))
    ).order_by('-total_amount')
    
    # Add data
    for row, member in enumerate(member_stats, 2):
        member_data = [
            member['member__member_id'],
            member['member__first_name'],
            member['member__last_name'],
            member['branch__code'],
            member['branch__name'],
            member['total_amount'],
            member['tithe_amount'],
            member['offering_amount'],
            member['contribution_count']
        ]
        
        for col, value in enumerate(member_data, 1):
            cell = ws_members.cell(row=row, column=col, value=value)
            if col >= 6 and col <= 8:  # Amount columns
                cell.number_format = '#,##0.00'
    
    # Auto-adjust column widths
    for column in ws_members.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws_members.column_dimensions[column_letter].width = adjusted_width
    
    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"member_contributions_{timezone.now().date()}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    wb.save(response)
    return response
