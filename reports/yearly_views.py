"""
Yearly Report Views - Professional, printable, audit-safe yearly financial reports

Reports are READ-ONLY and do NOT modify data.
Reports aggregate existing ledger-verified records from CLOSED MONTHS ONLY.
Reports must be exportable (PDF / CSV).

REPORT TYPES:
A) Mission Yearly Financial Report - Mission Admin & Auditors
B) Branch Yearly Financial Report - Branch Admin (own), Mission Admin, Auditors
C) Individual Member Yearly Contribution Statement - Member (self), Mission Admin, Auditors
"""

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import HttpResponse, JsonResponse
from django.db.models import Sum, Count, Q
from django.utils import timezone
from datetime import date, datetime
from decimal import Decimal
import calendar
import json

from core.models import Branch, Area, District, MonthlyClose, SiteSettings
from core.financial_helpers import (
    get_branch_balance, get_expected_mission_due, get_mission_income,
    get_mission_obligations, get_local_only_contributions
)
from contributions.models import Contribution, ContributionType, Remittance
from expenditure.models import Expenditure, ExpenditureCategory
from accounts.models import User


def get_unclosed_months_in_range(branch, year):
    """
    Check which months in a year are not closed for a branch.
    Returns list of unclosed month numbers and list of closed month numbers.
    """
    today = date.today()
    current_year = today.year
    current_month = today.month
    
    # Determine which months to check (don't check future months)
    if year < current_year:
        months_to_check = range(1, 13)
    elif year == current_year:
        months_to_check = range(1, current_month + 1)
    else:
        # Future year - no months to check
        return list(range(1, 13)), []
    
    closed_months = MonthlyClose.objects.filter(
        branch=branch,
        year=year,
        is_closed=True
    ).values_list('month', flat=True)
    
    closed_list = list(closed_months)
    unclosed_list = [m for m in months_to_check if m not in closed_list]
    
    return unclosed_list, closed_list


def get_all_branches_unclosed_months(year):
    """
    Get unclosed months across all branches for mission-level reports.
    Returns dict of branch_id -> list of unclosed months.
    """
    today = date.today()
    current_year = today.year
    current_month = today.month
    
    if year < current_year:
        months_to_check = list(range(1, 13))
    elif year == current_year:
        months_to_check = list(range(1, current_month + 1))
    else:
        return {}
    
    branches = Branch.objects.filter(is_active=True)
    unclosed_by_branch = {}
    
    for branch in branches:
        closed_months = MonthlyClose.objects.filter(
            branch=branch,
            year=year,
            is_closed=True
        ).values_list('month', flat=True)
        
        closed_list = list(closed_months)
        unclosed_list = [m for m in months_to_check if m not in closed_list]
        
        if unclosed_list:
            unclosed_by_branch[str(branch.id)] = {
                'branch_name': branch.name,
                'unclosed_months': unclosed_list
            }
    
    return unclosed_by_branch


@login_required
def yearly_reports_index(request):
    """Yearly reports dashboard - entry point for all yearly reports."""
    if not (request.user.is_mission_admin or request.user.is_auditor or 
            request.user.is_branch_executive or request.user.is_area_executive or 
            request.user.is_district_executive or request.user.is_pastor or
            request.user.role == 'member'):
        messages.error(request, 'Access denied.')
        return redirect('core:dashboard')
    
    today = date.today()
    available_years = list(range(today.year - 5, today.year + 1))
    
    context = {
        'available_years': available_years,
        'current_year': today.year,
        'user': request.user,
    }
    
    return render(request, 'reports/yearly_reports_index.html', context)


@login_required
def mission_yearly_report(request):
    """
    Mission Yearly Financial Report
    
    Access: Mission Admin, Auditors (read-only)
    
    Data Included:
    - Mission CASH income (from verified remittances)
    - Mission-level contribution types
    - Branch remittances received
    - Mission expenditures
    - Adjustments (if any)
    """
    if not (request.user.is_mission_admin or request.user.is_auditor):
        messages.error(request, 'Access denied. Mission Admin or Auditor role required.')
        return redirect('core:dashboard')
    
    # Get year parameter
    year = int(request.GET.get('year', date.today().year))
    proceed_partial = request.GET.get('proceed_partial', 'false') == 'true'
    
    # Date range for the year
    start_date = date(year, 1, 1)
    end_date = date(year, 12, 31)
    
    # Check for unclosed months across all branches
    unclosed_by_branch = get_all_branches_unclosed_months(year)
    has_unclosed_months = len(unclosed_by_branch) > 0
    
    # Get site settings
    site_settings = SiteSettings.get_settings()
    
    # ============ MISSION INCOME ============
    # CORRECT: Mission income comes ONLY from verified remittances
    mission_income_data = get_mission_income(start_date, end_date)
    total_remittances_received = mission_income_data['total_income']
    remittance_count = mission_income_data['remittance_count']
    
    # Branch-wise remittance breakdown
    remittances = Remittance.objects.filter(
        status__in=['verified', 'sent'],
        year=year
    ).select_related('branch')
    
    branch_remittances = remittances.values(
        'branch__name', 'branch__id', 'branch__code'
    ).annotate(
        total=Sum('amount_sent'),
        count=Count('id')
    ).order_by('-total')
    
    # Mission-level contributions (donations directly to mission)
    mission_contributions = Contribution.objects.filter(
        date__gte=start_date,
        date__lte=end_date,
        branch__isnull=True  # Mission-level only
    )
    
    mission_contrib_by_type = mission_contributions.values(
        'contribution_type__name'
    ).annotate(
        total=Sum('amount'),
        count=Count('id')
    ).order_by('-total')
    
    total_mission_contributions = mission_contributions.aggregate(
        total=Sum('amount')
    )['total'] or Decimal('0.00')
    
    total_mission_income = total_remittances_received + total_mission_contributions
    
    # ============ MISSION EXPENDITURES ============
    mission_expenditures = Expenditure.objects.filter(
        level='mission',
        date__gte=start_date,
        date__lte=end_date,
        status__in=['approved', 'paid']
    )
    
    total_mission_expenditures = mission_expenditures.aggregate(
        total=Sum('amount')
    )['total'] or Decimal('0.00')
    
    expenditures_by_category = mission_expenditures.values(
        'category__name'
    ).annotate(
        total=Sum('amount'),
        count=Count('id')
    ).order_by('-total')
    
    # ============ NET FINANCIAL POSITION ============
    net_position = total_mission_income - total_mission_expenditures
    
    # ============ OUTSTANDING REMITTANCES ============
    # Money owed but not yet received
    total_obligations = get_mission_obligations(start_date, end_date)
    outstanding_remittances = total_obligations - total_remittances_received
    
    # ============ DISCLOSURES ============
    # Get reopened months (months that were closed then reopened)
    # This would require tracking in MonthlyClose model - for now, we note if any months are unclosed
    
    # Available years for filter
    today = date.today()
    available_years = list(range(today.year - 5, today.year + 1))
    
    context = {
        'year': year,
        'start_date': start_date,
        'end_date': end_date,
        'generated_date': today,
        'site_settings': site_settings,
        
        # Unclosed months warning
        'has_unclosed_months': has_unclosed_months,
        'unclosed_by_branch': unclosed_by_branch,
        'proceed_partial': proceed_partial,
        
        # Income
        'total_remittances_received': total_remittances_received,
        'remittance_count': remittance_count,
        'branch_remittances': branch_remittances,
        'mission_contrib_by_type': mission_contrib_by_type,
        'total_mission_contributions': total_mission_contributions,
        'total_mission_income': total_mission_income,
        
        # Expenditures
        'total_mission_expenditures': total_mission_expenditures,
        'expenditures_by_category': expenditures_by_category,
        
        # Net Position
        'net_position': net_position,
        'outstanding_remittances': outstanding_remittances,
        
        # Filters
        'available_years': available_years,
    }
    
    return render(request, 'reports/mission_yearly_report.html', context)


@login_required
def branch_yearly_report(request):
    """
    Branch Yearly Financial Report
    
    Access: Branch Admin (own branch only), Mission Admin (read-only), Auditors
    
    Data Included:
    - Branch contributions
    - Branch expenditures
    - Mission share collected
    - Mission share remitted
    - Outstanding balances
    """
    # Determine access and branch
    branch_id = request.GET.get('branch')
    year = int(request.GET.get('year', date.today().year))
    proceed_partial = request.GET.get('proceed_partial', 'false') == 'true'
    
    if request.user.is_mission_admin or request.user.is_auditor:
        # Can view any branch
        if branch_id:
            branch = get_object_or_404(Branch, pk=branch_id)
        else:
            branch = Branch.objects.filter(is_active=True).first()
        branches = Branch.objects.filter(is_active=True).order_by('name')
    elif request.user.is_area_executive:
        # Can view branches in their area
        if branch_id:
            branch = get_object_or_404(Branch, pk=branch_id, district__area=request.user.managed_area)
        else:
            branch = Branch.objects.filter(district__area=request.user.managed_area, is_active=True).first()
        branches = Branch.objects.filter(district__area=request.user.managed_area, is_active=True).order_by('name')
        if not branch:
            messages.error(request, 'No branches found in your area.')
            return redirect('core:dashboard')
    elif request.user.is_district_executive:
        # Can view branches in their district
        if branch_id:
            branch = get_object_or_404(Branch, pk=branch_id, district=request.user.managed_district)
        else:
            branch = Branch.objects.filter(district=request.user.managed_district, is_active=True).first()
        branches = Branch.objects.filter(district=request.user.managed_district, is_active=True).order_by('name')
        if not branch:
            messages.error(request, 'No branches found in your district.')
            return redirect('core:dashboard')
    elif request.user.is_branch_executive or request.user.is_pastor:
        # Can only view own branch
        branch = request.user.branch
        branches = Branch.objects.filter(id=branch.id) if branch else Branch.objects.none()
        if not branch:
            messages.error(request, 'No branch assigned to your account.')
            return redirect('core:dashboard')
    else:
        messages.error(request, 'Access denied.')
        return redirect('core:dashboard')
    
    # Date range for the year
    start_date = date(year, 1, 1)
    end_date = date(year, 12, 31)
    
    # Check for unclosed months
    unclosed_months, closed_months = get_unclosed_months_in_range(branch, year)
    has_unclosed_months = len(unclosed_months) > 0
    
    # Get site settings
    site_settings = SiteSettings.get_settings()
    
    # ============ CONTRIBUTIONS ============
    contributions = Contribution.objects.filter(
        branch=branch,
        date__gte=start_date,
        date__lte=end_date
    )
    
    total_contributions = contributions.aggregate(
        total=Sum('amount')
    )['total'] or Decimal('0.00')
    
    contributions_by_type = contributions.values(
        'contribution_type__name', 'contribution_type__mission_percentage'
    ).annotate(
        total=Sum('amount'),
        count=Count('id')
    ).order_by('-total')
    
    # Monthly breakdown
    monthly_contributions = []
    for month in range(1, 13):
        month_total = contributions.filter(date__month=month).aggregate(
            total=Sum('amount')
        )['total'] or Decimal('0.00')
        monthly_contributions.append({
            'month': month,
            'month_name': calendar.month_name[month],
            'total': month_total,
            'is_closed': month in closed_months
        })
    
    # ============ EXPENDITURES ============
    expenditures = Expenditure.objects.filter(
        branch=branch,
        date__gte=start_date,
        date__lte=end_date,
        status__in=['approved', 'paid']
    )
    
    total_expenditures = expenditures.aggregate(
        total=Sum('amount')
    )['total'] or Decimal('0.00')
    
    expenditures_by_category = expenditures.values(
        'category__name'
    ).annotate(
        total=Sum('amount'),
        count=Count('id')
    ).order_by('-total')
    
    # ============ MISSION SHARE ============
    # Expected mission share (obligation)
    expected_mission_due = get_expected_mission_due(branch, start_date, end_date)
    
    # Actual remittances sent
    remittances = Remittance.objects.filter(
        branch=branch,
        year=year,
        status__in=['verified', 'sent']
    )
    
    total_remitted = remittances.aggregate(
        total=Sum('amount_sent')
    )['total'] or Decimal('0.00')
    
    outstanding_to_mission = expected_mission_due - total_remitted
    if outstanding_to_mission < 0:
        outstanding_to_mission = Decimal('0.00')
    
    # ============ FINANCIAL POSITION ============
    branch_balance_data = get_branch_balance(branch, start_date, end_date)
    branch_balance = branch_balance_data['branch_balance']
    
    # Opening balance (would need previous year's closing balance)
    # For now, we calculate based on current year data
    opening_balance = Decimal('0.00')  # TODO: Get from previous year's closing
    closing_balance = branch_balance
    
    # Available years and branches for filter
    today = date.today()
    available_years = list(range(today.year - 5, today.year + 1))
    
    context = {
        'branch': branch,
        'branches': branches,
        'year': year,
        'start_date': start_date,
        'end_date': end_date,
        'generated_date': today,
        'site_settings': site_settings,
        
        # Unclosed months warning
        'has_unclosed_months': has_unclosed_months,
        'unclosed_months': unclosed_months,
        'closed_months': closed_months,
        'proceed_partial': proceed_partial,
        
        # Contributions
        'total_contributions': total_contributions,
        'contributions_by_type': contributions_by_type,
        'monthly_contributions': monthly_contributions,
        
        # Expenditures
        'total_expenditures': total_expenditures,
        'expenditures_by_category': expenditures_by_category,
        
        # Mission Share
        'expected_mission_due': expected_mission_due,
        'total_remitted': total_remitted,
        'outstanding_to_mission': outstanding_to_mission,
        
        # Financial Position
        'opening_balance': opening_balance,
        'closing_balance': closing_balance,
        
        # Filters
        'available_years': available_years,
    }
    
    return render(request, 'reports/branch_yearly_report.html', context)


@login_required
def member_yearly_statement(request):
    """
    Individual Member Yearly Contribution Statement
    
    Access: Member (self only), Mission Admin (authorized), Auditors (read-only)
    
    Required Sections:
    - Member details
    - Contribution breakdown (Local + Mission donations)
    - Monthly totals (optional)
    - Declaration statement
    """
    member_id = request.GET.get('member')
    year = int(request.GET.get('year', date.today().year))
    
    # Get hierarchy filters
    area_id = request.GET.get('area')
    district_id = request.GET.get('district')
    branch_id = request.GET.get('branch')
    
@login_required
def member_yearly_statement(request):
    """
    Individual Member Yearly Contribution Statement
    
    Access: Member (self only), Mission Admin (authorized), Auditors (read-only)
    
    Required Sections:
    - Member details
    - Contribution breakdown (Local + Mission donations)
    - Monthly totals (optional)
    - Declaration statement
    """
    member_id = request.GET.get('member')
    year = int(request.GET.get('year', date.today().year))
    
    # Get hierarchy filters
    area_id = request.GET.get('area')
    district_id = request.GET.get('district')
    branch_id = request.GET.get('branch')
    
    # Determine access
    if request.user.is_mission_admin or request.user.is_auditor:
        # Can view any member
        if member_id:
            try:
                member = User.objects.select_related('branch', 'branch__district', 'branch__district__area').get(pk=member_id)
            except User.DoesNotExist:
                messages.error(request, 'Member not found.')
                return redirect('reports:yearly_reports_index')
        else:
            member = None
        
        # Get hierarchical data for filters
        areas = Area.objects.all().order_by('name')
        districts = District.objects.all().select_related('area').order_by('area__name', 'name')
        branches = Branch.objects.all().select_related('district', 'district__area').order_by('district__area__name', 'district__name', 'name')
        
        # Filter members based on hierarchy
        members_query = User.objects.filter(is_active=True, role='member').select_related('branch', 'branch__district', 'branch__district__area')
        
        if area_id:
            members_query = members_query.filter(branch__district__area_id=area_id)
            districts = districts.filter(area_id=area_id)
            branches = branches.filter(district__area_id=area_id)
        
        if district_id:
            members_query = members_query.filter(branch__district_id=district_id)
            branches = branches.filter(district_id=district_id)
        
        if branch_id:
            members_query = members_query.filter(branch_id=branch_id)
        
        members = members_query.order_by('branch__district__area__name', 'branch__district__name', 'branch__name', 'last_name', 'first_name')
        
    elif request.user.is_area_executive:
        # Can view members in their area
        if member_id:
            try:
                member = User.objects.select_related('branch', 'branch__district', 'branch__district__area').get(
                    pk=member_id, 
                    branch__district__area=request.user.managed_area
                )
            except User.DoesNotExist:
                messages.error(request, 'Member not found in your area.')
                return redirect('reports:yearly_reports_index')
        else:
            member = None
        
        # Get hierarchical data for their area
        areas = Area.objects.filter(pk=request.user.managed_area.pk).order_by('name')
        districts = District.objects.filter(area=request.user.managed_area).select_related('area').order_by('name')
        branches = Branch.objects.filter(district__area=request.user.managed_area).select_related('district', 'district__area').order_by('district__name', 'name')
        
        # Filter members based on hierarchy
        members_query = User.objects.filter(is_active=True, role='member', branch__district__area=request.user.managed_area).select_related('branch', 'branch__district', 'branch__district__area')
        
        if district_id:
            members_query = members_query.filter(branch__district_id=district_id)
            branches = branches.filter(district_id=district_id)
        
        if branch_id:
            members_query = members_query.filter(branch_id=branch_id)
        
        members = members_query.order_by('branch__district__name', 'branch__name', 'last_name', 'first_name')
        
    elif request.user.is_district_executive:
        # Can view members in their district
        if member_id:
            try:
                member = User.objects.select_related('branch', 'branch__district', 'branch__district__area').get(
                    pk=member_id, 
                    branch__district=request.user.managed_district
                )
            except User.DoesNotExist:
                messages.error(request, 'Member not found in your district.')
                return redirect('reports:yearly_reports_index')
        else:
            member = None
        
        # Get hierarchical data for their district
        areas = Area.objects.filter(pk=request.user.managed_district.area.pk).order_by('name')
        districts = District.objects.filter(pk=request.user.managed_district.pk).select_related('area').order_by('name')
        branches = Branch.objects.filter(district=request.user.managed_district).select_related('district', 'district__area').order_by('name')
        
        # Filter members based on hierarchy
        members_query = User.objects.filter(is_active=True, role='member', branch__district=request.user.managed_district).select_related('branch', 'branch__district', 'branch__district__area')
        
        if branch_id:
            members_query = members_query.filter(branch_id=branch_id)
        
        members = members_query.order_by('branch__name', 'last_name', 'first_name')
        
    else:
        # Regular member - can only view own statement
        member = request.user
        members = User.objects.filter(pk=request.user.pk)
        areas = Area.objects.none()
        districts = District.objects.none()
        branches = Branch.objects.none()
    
    # Get site settings
    site_settings = SiteSettings.get_settings()
    
    # Date range for the year
    start_date = date(year, 1, 1)
    end_date = date(year, 12, 31)
    today = date.today()
    
    # Initialize data
    contributions_by_type = []
    monthly_totals = []
    total_contributions = Decimal('0.00')
    
    if member:
        # Get contributions for this member
        contributions = Contribution.objects.filter(
            member=member,
            date__gte=start_date,
            date__lte=end_date
        ).select_related('contribution_type', 'branch')
        
        total_contributions = contributions.aggregate(
            total=Sum('amount')
        )['total'] or Decimal('0.00')
        
        # By contribution type
        contributions_by_type = contributions.values(
            'contribution_type__name'
        ).annotate(
            total=Sum('amount'),
            count=Count('id')
        ).order_by('-total')
        
        # Monthly breakdown
        for month in range(1, 13):
            month_total = contributions.filter(date__month=month).aggregate(
                total=Sum('amount')
            )['total'] or Decimal('0.00')
            monthly_totals.append({
                'month': month,
                'month_name': calendar.month_name[month],
                'total': month_total
            })
    
    # Available years for filter
    available_years = list(range(today.year - 5, today.year + 1))
    
    context = {
        'member': member,
        'members': members,
        'year': year,
        'start_date': start_date,
        'end_date': end_date,
        'generated_date': today,
        'site_settings': site_settings,
        
        # Hierarchical data
        'areas': areas,
        'districts': districts,
        'branches': branches,
        
        # Contributions
        'total_contributions': total_contributions,
        'contributions_by_type': contributions_by_type,
        'monthly_totals': monthly_totals,
        
        # Filters
        'available_years': available_years,
    }
    
    return render(request, 'reports/member_yearly_statement.html', context)


def export_mission_yearly_pdf(request):
    """Export Mission Yearly Report as PDF."""
    from django.template.loader import render_to_string
    
    # Check permissions first
    if not (request.user.is_mission_admin or request.user.is_auditor):
        messages.error(request, 'Access denied.')
        return redirect('core:dashboard')
    
    year = int(request.GET.get('year', date.today().year))
    
    # Get the same data as the view
    start_date = date(year, 1, 1)
    end_date = date(year, 12, 31)
    site_settings = SiteSettings.get_settings()
    
    # Mission income
    mission_income_data = get_mission_income(start_date, end_date)
    total_remittances_received = mission_income_data['total_income']
    
    # Branch remittances
    remittances = Remittance.objects.filter(
        status__in=['verified', 'sent'],
        year=year
    ).select_related('branch')
    
    branch_remittances = remittances.values(
        'branch__name', 'branch__code'
    ).annotate(
        total=Sum('amount_sent'),
        count=Count('id')
    ).order_by('-total')
    
    # Mission contributions
    mission_contributions = Contribution.objects.filter(
        date__gte=start_date,
        date__lte=end_date,
        branch__isnull=True
    )
    total_mission_contributions = mission_contributions.aggregate(
        total=Sum('amount')
    )['total'] or Decimal('0.00')
    
    total_mission_income = total_remittances_received + total_mission_contributions
    
    # Mission expenditures
    mission_expenditures = Expenditure.objects.filter(
        level='mission',
        date__gte=start_date,
        date__lte=end_date,
        status__in=['approved', 'paid']
    )
    total_mission_expenditures = mission_expenditures.aggregate(
        total=Sum('amount')
    )['total'] or Decimal('0.00')
    
    expenditures_by_category = mission_expenditures.values(
        'category__name'
    ).annotate(
        total=Sum('amount')
    ).order_by('-total')
    
    net_position = total_mission_income - total_mission_expenditures
    
    context = {
        'year': year,
        'generated_date': date.today(),
        'site_settings': site_settings,
        'total_remittances_received': total_remittances_received,
        'branch_remittances': branch_remittances,
        'total_mission_contributions': total_mission_contributions,
        'total_mission_income': total_mission_income,
        'total_mission_expenditures': total_mission_expenditures,
        'expenditures_by_category': expenditures_by_category,
        'net_position': net_position,
    }
    
    html_string = render_to_string('reports/mission_yearly_report_pdf.html', context)
    
    try:
        # Import WeasyPrint here to catch library loading errors
        from weasyprint import HTML
        html = HTML(string=html_string)
        pdf = html.write_pdf()
        
        response = HttpResponse(pdf, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="mission_yearly_report_{year}.pdf"'
        return response
    except Exception as e:
        messages.error(request, f'PDF generation failed: {str(e)}. Please try the Excel export instead.')
        return redirect('reports:mission_yearly_report')


@login_required
def export_mission_yearly_excel(request):
    """Export Mission Yearly Report to Excel."""
    year = int(request.GET.get('year', date.today().year))
    
    # Check permissions
    if not (request.user.is_mission_admin or request.user.is_auditor):
        messages.error(request, 'Access denied.')
        return redirect('core:dashboard')
    
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        messages.error(request, 'Excel export requires openpyxl package to be installed.')
        return redirect('reports:mission_yearly_report')
    
    # Get the same data as the view
    start_date = date(year, 1, 1)
    end_date = date(year, 12, 31)
    site_settings = SiteSettings.get_settings()
    
    # Mission income
    mission_income_data = get_mission_income(start_date, end_date)
    total_remittances_received = mission_income_data['total_income']
    
    # Branch remittances
    remittances = Remittance.objects.filter(
        status__in=['verified', 'sent'],
        year=year
    ).select_related('branch')
    
    branch_remittances = remittances.values(
        'branch__name', 'branch__code'
    ).annotate(
        total=Sum('amount_sent'),
        count=Count('id')
    ).order_by('-total')
    
    # Mission contributions
    mission_contributions = Contribution.objects.filter(
        date__gte=start_date,
        date__lte=end_date,
        branch__isnull=True
    )
    total_mission_contributions = mission_contributions.aggregate(
        total=Sum('amount')
    )['total'] or Decimal('0.00')
    
    total_mission_income = total_remittances_received + total_mission_contributions
    
    # Mission expenditures
    mission_expenditures = Expenditure.objects.filter(
        level='mission',
        date__gte=start_date,
        date__lte=end_date,
        status__in=['approved', 'paid']
    )
    total_mission_expenditures = mission_expenditures.aggregate(
        total=Sum('amount')
    )['total'] or Decimal('0.00')
    
    expenditures_by_category = mission_expenditures.values(
        'category__name'
    ).annotate(
        total=Sum('amount'),
        count=Count('id')
    ).order_by('-total')
    
    net_position = total_mission_income - total_mission_expenditures
    
    # Create workbook
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    
    # Define styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")  # Indigo
    header_alignment = Alignment(horizontal="center", vertical="center")
    border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                   top=Side(style='thin'), bottom=Side(style='thin'))
    
    # Summary sheet
    ws_summary = wb.create_sheet("Summary")
    
    # Report info
    ws_summary.merge_cells('A1:B1')
    ws_summary['A1'] = f"Mission Yearly Financial Report - {year}"
    ws_summary['A1'].font = Font(bold=True, size=14)
    ws_summary['A1'].alignment = Alignment(horizontal="center")
    
    ws_summary['A3'] = "Report Period:"
    ws_summary['B3'] = f"{start_date.strftime('%B %d, %Y')} - {end_date.strftime('%B %d, %Y')}"
    ws_summary['A4'] = "Generated:"
    ws_summary['B4'] = date.today().strftime('%B %d, %Y')
    
    # Financial Summary
    ws_summary['A6'] = "Financial Summary"
    ws_summary['A6'].font = Font(bold=True, size=12)
    
    summary_data = [
        ("Total Remittances Received", total_remittances_received),
        ("Mission Contributions", total_mission_contributions),
        ("Total Mission Income", total_mission_income),
        ("Total Mission Expenditure", total_mission_expenditures),
        ("Net Position", net_position)
    ]
    
    row = 7
    for label, amount in summary_data:
        ws_summary[f'A{row}'] = label
        ws_summary[f'B{row}'] = float(amount)
        ws_summary[f'B{row}'].number_format = '#,##0.00'
        row += 1
    
    # Branch Remittances sheet
    ws_remittances = wb.create_sheet("Branch Remittances")
    headers = ["Branch Code", "Branch Name", "Amount Sent", "Count"]
    
    for col, header in enumerate(headers, 1):
        cell = ws_remittances.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    
    row = 2
    for item in branch_remittances:
        ws_remittances[f'A{row}'] = item['branch__code']
        ws_remittances[f'B{row}'] = item['branch__name']
        ws_remittances[f'C{row}'] = float(item['total'])
        ws_remittances[f'C{row}'].number_format = '#,##0.00'
        ws_remittances[f'D{row}'] = item['count']
        row += 1
    
    # Total row
    ws_remittances[f'A{row}'] = "Total"
    ws_remittances[f'A{row}'].font = Font(bold=True)
    ws_remittances[f'B{row}'].font = Font(bold=True)
    ws_remittances[f'C{row}'] = float(total_remittances_received)
    ws_remittances[f'C{row}'].font = Font(bold=True)
    ws_remittances[f'C{row}'].number_format = '#,##0.00'
    ws_remittances[f'D{row}'] = branch_remittances.count()
    ws_remittances[f'D{row}'].font = Font(bold=True)
    
    # Mission Expenditures sheet
    ws_expenditures = wb.create_sheet("Mission Expenditures")
    headers = ["Category", "Count", "Total Amount"]
    
    for col, header in enumerate(headers, 1):
        cell = ws_expenditures.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    
    row = 2
    for item in expenditures_by_category:
        ws_expenditures[f'A{row}'] = item['category__name'] or "Uncategorized"
        ws_expenditures[f'B{row}'] = item['count']
        ws_expenditures[f'C{row}'] = float(item['total'])
        ws_expenditures[f'C{row}'].number_format = '#,##0.00'
        row += 1
    
    # Total row
    ws_expenditures[f'A{row}'] = "Total"
    ws_expenditures[f'A{row}'].font = Font(bold=True)
    ws_expenditures[f'B{row}'] = mission_expenditures.count()
    ws_expenditures[f'B{row}'].font = Font(bold=True)
    ws_expenditures[f'C{row}'] = float(total_mission_expenditures)
    ws_expenditures[f'C{row}'].font = Font(bold=True)
    ws_expenditures[f'C{row}'].number_format = '#,##0.00'
    
    # Auto-adjust column widths
    for ws in [ws_summary, ws_remittances, ws_expenditures]:
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save to response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="mission_yearly_report_{year}.xlsx"'
    
    wb.save(response)
    return response


@login_required
def export_branch_yearly_pdf(request):
    """Export Branch Yearly Report as PDF."""
    from django.template.loader import render_to_string
    
    branch_id = request.GET.get('branch')
    year = int(request.GET.get('year', date.today().year))
    
    # Determine access and branch
    if request.user.is_mission_admin or request.user.is_auditor:
        if branch_id:
            branch = get_object_or_404(Branch, pk=branch_id)
        else:
            messages.error(request, 'Branch not specified.')
            return redirect('reports:branch_yearly_report')
    elif request.user.is_branch_executive or request.user.is_pastor:
        branch = request.user.branch
        if not branch:
            messages.error(request, 'No branch assigned.')
            return redirect('core:dashboard')
    else:
        messages.error(request, 'Access denied.')
        return redirect('core:dashboard')
    
    start_date = date(year, 1, 1)
    end_date = date(year, 12, 31)
    site_settings = SiteSettings.get_settings()
    
    # Get data
    contributions = Contribution.objects.filter(
        branch=branch,
        date__gte=start_date,
        date__lte=end_date
    )
    total_contributions = contributions.aggregate(total=Sum('amount'))['total'] or Decimal('0.00')
    
    contributions_by_type = contributions.values(
        'contribution_type__name'
    ).annotate(
        total=Sum('amount')
    ).order_by('-total')
    
    expenditures = Expenditure.objects.filter(
        branch=branch,
        date__gte=start_date,
        date__lte=end_date,
        status__in=['approved', 'paid']
    )
    total_expenditures = expenditures.aggregate(total=Sum('amount'))['total'] or Decimal('0.00')
    
    expected_mission_due = get_expected_mission_due(branch, start_date, end_date)
    
    remittances = Remittance.objects.filter(
        branch=branch,
        year=year,
        status__in=['verified', 'sent']
    )
    total_remitted = remittances.aggregate(total=Sum('amount_sent'))['total'] or Decimal('0.00')
    
    outstanding_to_mission = expected_mission_due - total_remitted
    if outstanding_to_mission < 0:
        outstanding_to_mission = Decimal('0.00')
    
    branch_balance_data = get_branch_balance(branch, start_date, end_date)
    closing_balance = branch_balance_data['branch_balance']
    
    context = {
        'branch': branch,
        'year': year,
        'generated_date': date.today(),
        'site_settings': site_settings,
        'total_contributions': total_contributions,
        'contributions_by_type': contributions_by_type,
        'total_expenditures': total_expenditures,
        'expected_mission_due': expected_mission_due,
        'total_remitted': total_remitted,
        'outstanding_to_mission': outstanding_to_mission,
        'closing_balance': closing_balance,
    }
    
    html_string = render_to_string('reports/branch_yearly_report_pdf.html', context)
    
    try:
        # Import WeasyPrint here to catch library loading errors
        from weasyprint import HTML
        html = HTML(string=html_string)
        pdf = html.write_pdf()
        
        response = HttpResponse(pdf, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="branch_yearly_report_{branch.code}_{year}.pdf"'
        return response
    except Exception as e:
        messages.error(request, f'PDF generation failed: {str(e)}. Please try the Excel export instead.')
        return redirect('reports:branch_yearly_report')


@login_required
def export_member_yearly_pdf(request):
    """Export Member Yearly Statement as PDF."""
    from django.template.loader import render_to_string
    
    member_id = request.GET.get('member')
    year = int(request.GET.get('year', date.today().year))
    
    # Determine access
    if request.user.is_mission_admin or request.user.is_auditor:
        if member_id:
            try:
                member = User.objects.select_related('branch', 'branch__district').get(pk=member_id)
            except User.DoesNotExist:
                messages.error(request, 'Member not found.')
                return redirect('reports:member_yearly_statement')
        else:
            messages.error(request, 'Member not specified.')
            return redirect('reports:member_yearly_statement')
    else:
        member = request.user
    
    start_date = date(year, 1, 1)
    end_date = date(year, 12, 31)
    site_settings = SiteSettings.get_settings()
    
    contributions = Contribution.objects.filter(
        member=member,
        date__gte=start_date,
        date__lte=end_date
    ).select_related('contribution_type')
    
    total_contributions = contributions.aggregate(total=Sum('amount'))['total'] or Decimal('0.00')
    
    contributions_by_type = contributions.values(
        'contribution_type__name'
    ).annotate(
        total=Sum('amount'),
        count=Count('id')
    ).order_by('-total')
    
    monthly_totals = []
    for month in range(1, 13):
        month_total = contributions.filter(date__month=month).aggregate(
            total=Sum('amount')
        )['total'] or Decimal('0.00')
        monthly_totals.append({
            'month_name': calendar.month_name[month],
            'total': month_total
        })
    
    context = {
        'member': member,
        'year': year,
        'generated_date': date.today(),
        'site_settings': site_settings,
        'total_contributions': total_contributions,
        'contributions_by_type': contributions_by_type,
        'monthly_totals': monthly_totals,
    }
    
    html_string = render_to_string('reports/member_yearly_statement_pdf.html', context)
    
    try:
        # Import WeasyPrint here to catch library loading errors
        from weasyprint import HTML
        html = HTML(string=html_string)
        pdf = html.write_pdf()
        
        response = HttpResponse(pdf, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="contribution_statement_{member.member_id}_{year}.pdf"'
        return response
    except Exception as e:
        messages.error(request, f'PDF generation failed: {str(e)}. Please try again later or contact support.')
        return redirect('reports:member_yearly_statement')


@login_required
def export_member_yearly_excel(request):
    """Export Member Yearly Statement to Excel."""
    member_id = request.GET.get('member')
    year = int(request.GET.get('year', date.today().year))
    
    # Determine access
    if request.user.is_mission_admin or request.user.is_auditor:
        if member_id:
            try:
                member = User.objects.select_related('branch', 'branch__district').get(pk=member_id)
            except User.DoesNotExist:
                messages.error(request, 'Member not found.')
                return redirect('reports:member_yearly_statement')
        else:
            messages.error(request, 'Member not specified.')
            return redirect('reports:member_yearly_statement')
    else:
        member = request.user
    
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        messages.error(request, 'Excel export requires openpyxl package to be installed.')
        return redirect('reports:member_yearly_statement')
    
    # Date range for the year
    start_date = date(year, 1, 1)
    end_date = date(year, 12, 31)
    site_settings = SiteSettings.get_settings()
    
    # Get contributions for this member
    contributions = Contribution.objects.filter(
        member=member,
        date__gte=start_date,
        date__lte=end_date
    ).select_related('contribution_type', 'branch')
    
    total_contributions = contributions.aggregate(
        total=Sum('amount')
    )['total'] or Decimal('0.00')
    
    # By contribution type
    contributions_by_type = contributions.values(
        'contribution_type__name'
    ).annotate(
        total=Sum('amount'),
        count=Count('id')
    ).order_by('-total')
    
    # Monthly breakdown
    monthly_totals = []
    for month in range(1, 13):
        month_total = contributions.filter(date__month=month).aggregate(
            total=Sum('amount')
        )['total'] or Decimal('0.00')
        monthly_totals.append({
            'month': month,
            'month_name': calendar.month_name[month],
            'total': month_total
        })
    
    # Create workbook
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    
    # Define styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")  # Blue
    header_alignment = Alignment(horizontal="center", vertical="center")
    border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                   top=Side(style='thin'), bottom=Side(style='thin'))
    
    # Member Info sheet
    ws_member = wb.create_sheet("Member Information")
    
    # Member details
    ws_member.merge_cells('A1:C1')
    ws_member['A1'] = f"Member Contribution Statement - {year}"
    ws_member['A1'].font = Font(bold=True, size=14)
    ws_member['A1'].alignment = Alignment(horizontal="center")
    
    ws_member['A3'] = "Member Name:"
    ws_member['B3'] = member.get_full_name()
    ws_member['A4'] = "Member ID:"
    ws_member['B4'] = member.member_id or "N/A"
    ws_member['A5'] = "Branch:"
    ws_member['B5'] = member.branch.name if member.branch else "No Branch"
    ws_member['A6'] = "Report Period:"
    ws_member['B6'] = f"January 1, {year} - December 31, {year}"
    ws_member['A7'] = "Generated:"
    ws_member['B7'] = date.today().strftime('%B %d, %Y')
    
    # Total Contributions
    ws_member['A9'] = "Total Contributions for {year}"
    ws_member['A9'].font = Font(bold=True, size=12)
    ws_member['A10'] = site_settings.currency_symbol
    ws_member['B10'] = float(total_contributions)
    ws_member['B10'].font = Font(bold=True, size=14)
    ws_member['B10'].number_format = '#,##0.00'
    
    # Contributions by Type sheet
    ws_contributions = wb.create_sheet("Contributions by Type")
    headers = ["Contribution Type", "Count", "Total Amount"]
    
    for col, header in enumerate(headers, 1):
        cell = ws_contributions.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    
    row = 2
    for item in contributions_by_type:
        ws_contributions[f'A{row}'] = item['contribution_type__name']
        ws_contributions[f'B{row}'] = item['count']
        ws_contributions[f'C{row}'] = float(item['total'])
        ws_contributions[f'C{row}'].number_format = '#,##0.00'
        row += 1
    
    # Total row
    ws_contributions[f'A{row}'] = "Total"
    ws_contributions[f'A{row}'].font = Font(bold=True)
    ws_contributions[f'B{row}'] = contributions.count()
    ws_contributions[f'B{row}'].font = Font(bold=True)
    ws_contributions[f'C{row}'] = float(total_contributions)
    ws_contributions[f'C{row}'].font = Font(bold=True)
    ws_contributions[f'C{row}'].number_format = '#,##0.00'
    
    # Monthly Breakdown sheet
    ws_monthly = wb.create_sheet("Monthly Breakdown")
    headers = ["Month", "Contributions"]
    
    for col, header in enumerate(headers, 1):
        cell = ws_monthly.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    
    row = 2
    for month_data in monthly_totals:
        ws_monthly[f'A{row}'] = month_data['month_name']
        ws_monthly[f'B{row}'] = float(month_data['total'])
        ws_monthly[f'B{row}'].number_format = '#,##0.00'
        row += 1
    
    # Total row
    ws_monthly[f'A{row}'] = "Total"
    ws_monthly[f'A{row}'].font = Font(bold=True)
    ws_monthly[f'B{row}'] = float(total_contributions)
    ws_monthly[f'B{row}'].font = Font(bold=True)
    ws_monthly[f'B{row}'].number_format = '#,##0.00'
    
    # Declaration sheet
    ws_declaration = wb.create_sheet("Declaration")
    ws_declaration.merge_cells('A1:B1')
    ws_declaration['A1'] = "Declaration"
    ws_declaration['A1'].font = Font(bold=True, size=14)
    ws_declaration['A1'].alignment = Alignment(horizontal="center")
    
    ws_declaration['A3'] = "I hereby declare that the contributions listed above are accurate and complete"
    ws_declaration['A4'] = "to the best of my knowledge for the calendar year {year}."
    ws_declaration['A5'] = ""
    ws_declaration['A6'] = "_________________________"
    ws_declaration['A7'] = f"{member.get_full_name()}"
    ws_declaration['A8'] = f"Date: {date.today().strftime('%B %d, %Y')}"
    
    # Auto-adjust column widths
    for ws in [ws_member, ws_contributions, ws_monthly, ws_declaration]:
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save to response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="contribution_statement_{member.member_id or member.id}_{year}.xlsx"'
    
    wb.save(response)
    return response


@login_required
def export_branch_yearly_excel(request):
    """Export branch yearly financial report to Excel."""
    branch_id = request.GET.get('branch')
    year = int(request.GET.get('year', timezone.now().year))
    
    # Get branch and verify permissions
    branch = get_object_or_404(Branch, pk=branch_id)
    if not (request.user.is_mission_admin or request.user.is_auditor or 
            (request.user.is_branch_executive and request.user.branch == branch)):
        messages.error(request, 'Access denied.')
        return redirect('core:dashboard')
    
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        messages.error(request, 'Excel export requires openpyxl package to be installed.')
        return redirect('reports:branch_yearly_report')
    
    # Get the same data as the regular report
    start_date = date(year, 1, 1)
    end_date = date(year, 12, 31)
    site_settings = SiteSettings.get_settings()
    
    # Get unclosed months info
    unclosed_months, closed_months = get_unclosed_months_in_range(branch, year)
    has_unclosed_months = len(unclosed_months) > 0
    proceed_partial = request.GET.get('proceed_partial') == 'true'
    
    # Filter by closed months if specified
    if has_unclosed_months and not proceed_partial:
        messages.error(request, 'Some months are not closed. Please close all months or proceed with partial data.')
        return redirect('reports:branch_yearly_report')
    
    # Get contributions data
    contributions_filter = Q(branch=branch, date__gte=start_date, date__lte=end_date)
    if has_unclosed_months and proceed_partial:
        contributions_filter &= Q(date__month__in=closed_months)
    
    contributions = Contribution.objects.filter(contributions_filter).select_related('contribution_type')
    total_contributions = contributions.aggregate(total=Sum('amount'))['total'] or Decimal('0.00')
    
    contributions_by_type = contributions.values(
        'contribution_type__name'
    ).annotate(
        total=Sum('amount'),
        count=Count('id')
    ).order_by('-total')
    
    # Get expenditures data
    expenditures_filter = Q(branch=branch, date__gte=start_date, date__lte=end_date)
    if has_unclosed_months and proceed_partial:
        expenditures_filter &= Q(date__month__in=closed_months)
    
    expenditures = Expenditure.objects.filter(expenditures_filter).select_related('category')
    total_expenditures = expenditures.aggregate(total=Sum('amount'))['total'] or Decimal('0.00')
    
    expenditures_by_category = expenditures.values(
        'category__name'
    ).annotate(
        total=Sum('amount'),
        count=Count('id')
    ).order_by('-total')
    
    # Get remittances data
    remittances_filter = Q(branch=branch, date__gte=start_date, date__lte=end_date)
    if has_unclosed_months and proceed_partial:
        remittances_filter &= Q(date__month__in=closed_months)
    
    remittances = Remittance.objects.filter(remittances_filter)
    total_remitted = remittances.aggregate(total=Sum('amount'))['total'] or Decimal('0.00')
    
    # Calculate financial position
    opening_balance = get_branch_balance(branch, start_date - timezone.timedelta(days=1))
    expected_mission_due = get_expected_mission_due(branch, start_date, end_date, 
                                                   include_unclosed=not (has_unclosed_months and proceed_partial))
    outstanding_to_mission = expected_mission_due - total_remitted
    closing_balance = opening_balance + total_contributions - total_expenditures - total_remitted
    
    # Create workbook
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    
    # Define styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="6B46C1", end_color="6B46C1", fill_type="solid")  # Violet
    header_alignment = Alignment(horizontal="center", vertical="center")
    border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                   top=Side(style='thin'), bottom=Side(style='thin'))
    
    # Summary sheet
    ws_summary = wb.create_sheet("Summary")
    
    # Branch info
    ws_summary.merge_cells('A1:B1')
    ws_summary['A1'] = f"Branch Yearly Financial Report - {branch.name}"
    ws_summary['A1'].font = Font(bold=True, size=14)
    ws_summary['A1'].alignment = Alignment(horizontal="center")
    
    ws_summary['A3'] = "Branch:"
    ws_summary['B3'] = branch.name
    ws_summary['A4'] = "District:"
    ws_summary['B4'] = branch.district.name
    ws_summary['A5'] = "Area:"
    ws_summary['B5'] = branch.district.area.name
    ws_summary['A6'] = "Year:"
    ws_summary['B6'] = str(year)
    ws_summary['A7'] = "Report Period:"
    ws_summary['B7'] = f"{start_date.strftime('%B %d, %Y')} - {end_date.strftime('%B %d, %Y')}"
    
    # Financial Summary
    ws_summary['A9'] = "Financial Summary"
    ws_summary['A9'].font = Font(bold=True, size=12)
    
    summary_data = [
        ("Total Contributions", total_contributions),
        ("Total Expenditure", total_expenditures),
        ("Mission Share Remitted", total_remitted),
        ("Mission Share Due", expected_mission_due),
        ("Outstanding Balance", outstanding_to_mission),
        ("Branch Balance (Closing)", closing_balance)
    ]
    
    row = 10
    for label, amount in summary_data:
        ws_summary[f'A{row}'] = label
        ws_summary[f'B{row}'] = float(amount)
        ws_summary[f'B{row}'].number_format = '#,##0.00'
        row += 1
    
    # Contributions by Type sheet
    ws_contrib = wb.create_sheet("Contributions by Type")
    headers = ["Contribution Type", "Count", "Total Amount"]
    
    for col, header in enumerate(headers, 1):
        cell = ws_contrib.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    
    row = 2
    for item in contributions_by_type:
        ws_contrib[f'A{row}'] = item['contribution_type__name']
        ws_contrib[f'B{row}'] = item['count']
        ws_contrib[f'C{row}'] = float(item['total'])
        ws_contrib[f'C{row}'].number_format = '#,##0.00'
        row += 1
    
    # Total row
    ws_contrib[f'A{row}'] = "Total"
    ws_contrib[f'A{row}'].font = Font(bold=True)
    ws_contrib[f'B{row}'] = contributions.count()
    ws_contrib[f'B{row}'].font = Font(bold=True)
    ws_contrib[f'C{row}'] = float(total_contributions)
    ws_contrib[f'C{row}'].font = Font(bold=True)
    ws_contrib[f'C{row}'].number_format = '#,##0.00'
    
    # Expenditures by Category sheet
    ws_exp = wb.create_sheet("Expenditures by Category")
    headers = ["Category", "Count", "Total Amount"]
    
    for col, header in enumerate(headers, 1):
        cell = ws_exp.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    
    row = 2
    for item in expenditures_by_category:
        ws_exp[f'A{row}'] = item['category__name'] or "Uncategorized"
        ws_exp[f'B{row}'] = item['count']
        ws_exp[f'C{row}'] = float(item['total'])
        ws_exp[f'C{row}'].number_format = '#,##0.00'
        row += 1
    
    # Total row
    ws_exp[f'A{row}'] = "Total"
    ws_exp[f'A{row}'].font = Font(bold=True)
    ws_exp[f'B{row}'] = expenditures.count()
    ws_exp[f'B{row}'].font = Font(bold=True)
    ws_exp[f'C{row}'] = float(total_expenditures)
    ws_exp[f'C{row}'].font = Font(bold=True)
    ws_exp[f'C{row}'].number_format = '#,##0.00'
    
    # Monthly Breakdown sheet
    ws_monthly = wb.create_sheet("Monthly Breakdown")
    headers = ["Month", "Contributions", "Status"]
    
    for col, header in enumerate(headers, 1):
        cell = ws_monthly.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    
    row = 2
    for month in range(1, 13):
        month_name = calendar.month_name[month]
        month_contributions = contributions.filter(date__month=month).aggregate(total=Sum('amount'))['total'] or Decimal('0.00')
        is_closed = month in closed_months
        status = "Closed" if is_closed else "Open"
        
        ws_monthly[f'A{row}'] = month_name
        ws_monthly[f'B{row}'] = float(month_contributions)
        ws_monthly[f'B{row}'].number_format = '#,##0.00'
        ws_monthly[f'C{row}'] = status
        row += 1
    
    # Auto-adjust column widths
    for ws in [ws_summary, ws_contrib, ws_exp, ws_monthly]:
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save to response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="branch_yearly_report_{branch.name}_{year}.xlsx"'
    
    wb.save(response)
    return response
