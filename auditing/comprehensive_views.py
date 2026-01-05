"""
Comprehensive Auditor Views - Complete auditing and reporting system
"""

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import HttpResponse, JsonResponse
from django.db.models import Sum, Count, Q, F
from django.utils import timezone
from datetime import date, timedelta, datetime
from decimal import Decimal
import calendar
import csv

from core.models import Branch, District, Area, FiscalYear, MonthlyClose
from contributions.models import Contribution, Remittance, ContributionType
from expenditure.models import Expenditure, ExpenditureCategory
from payroll.models import PayrollRun, PaySlip
from accounts.models import User


@login_required
def auditor_dashboard(request):
    """Comprehensive auditor dashboard with all financial oversight."""
    if not (request.user.is_auditor or request.user.is_mission_admin):
        messages.error(request, 'Access denied. Auditor role required.')
        return redirect('core:dashboard')
    
    today = date.today()
    current_month = today.month
    current_year = today.year
    
    # Get filters
    area_id = request.GET.get('area')
    district_id = request.GET.get('district')
    branch_id = request.GET.get('branch')
    
    # Get fiscal year
    # DEPRECATED: Year-as-state architecture - Use date filtering only
    fiscal_year = None  # Not used anymore
    
    # Filter branches based on hierarchy
    branches = Branch.objects.filter(is_active=True)
    if area_id:
        branches = branches.filter(district__area_id=area_id)
    if district_id:
        branches = branches.filter(district_id=district_id)
    if branch_id:
        branches = branches.filter(id=branch_id)
    
    total_branches = branches.count()
    
    # Current month statistics
    month_start = date(current_year, current_month, 1)
    if current_month == 12:
        month_end = date(current_year + 1, 1, 1) - timedelta(days=1)
    else:
        month_end = date(current_year, current_month + 1, 1) - timedelta(days=1)
    
    # Contributions this month
    # DEPRECATED: Year-as-state architecture - Use date filtering only
    month_contributions = Contribution.objects.filter(
        branch__in=branches,
        date__gte=month_start,
        date__lte=month_end
    ).aggregate(
        total=Sum('amount'),
        count=Count('id')
    )
    
    # Expenditures this month
    # DEPRECATED: Year-as-state architecture - Use date filtering only
    month_expenditures = Expenditure.objects.filter(
        branch__in=branches,
        date__gte=month_start,
        date__lte=month_end
    ).aggregate(
        total=Sum('amount'),
        count=Count('id')
    )
    
    # Remittances pending
    pending_remittances = Remittance.objects.filter(
        branch__in=branches,
        status__in=['pending', 'sent']
    ).aggregate(
        total=Sum('amount_due'),
        count=Count('id')
    )
    
    # Closed months this year
    closed_months = MonthlyClose.objects.filter(
        year=current_year,
        is_closed=True
    ).count()
    
    # Top contributing branches this month
    # DEPRECATED: Year-as-state architecture - Use date filtering only
    top_branches = Contribution.objects.filter(
        branch__in=branches,
        date__gte=month_start,
        date__lte=month_end
    ).values('branch__name', 'branch__id').annotate(
        total=Sum('amount')
    ).order_by('-total')[:10]
    
    # Recent transactions (last 20)
    recent_contributions = Contribution.objects.filter(
        branch__in=branches
    ).select_related(
        'branch', 'contribution_type', 'created_by'
    ).order_by('-created_at')[:10]
    
    recent_expenditures = Expenditure.objects.filter(
        branch__in=branches
    ).select_related(
        'branch', 'category', 'created_by'
    ).order_by('-created_at')[:10]
    
    # Branch performance data
    branch_performance = []
    for branch in branches[:20]:  # Limit to 20 for dashboard
        branch_contribs = Contribution.objects.filter(
            branch=branch,
            date__gte=month_start,
            date__lte=month_end
        ).aggregate(total=Sum('amount'))['total'] or Decimal('0.00')
        
        branch_expends = Expenditure.objects.filter(
            branch=branch,
            date__gte=month_start,
            date__lte=month_end
        ).aggregate(total=Sum('amount'))['total'] or Decimal('0.00')
        
        # Calculate mission due (10% of tithe)
        tithe_amount = Contribution.objects.filter(
            branch=branch,
            date__gte=month_start,
            date__lte=month_end,
            contribution_type__category='tithe'
        ).aggregate(total=Sum('amount'))['total'] or Decimal('0.00')
        
        mission_due = tithe_amount * Decimal('0.10')
        
        # Get remittance status
        remittance = Remittance.objects.filter(
            branch=branch,
            year=current_year,
            month=current_month
        ).first()
        
        remittance_status = 'pending'
        if remittance:
            remittance_status = remittance.status
        
        branch_performance.append({
            'name': branch.name,
            'total_contributions': branch_contribs,
            'mission_due': mission_due,
            'total_expenditures': branch_expends,
            'net_balance': branch_contribs - branch_expends,
            'remittance_status': remittance_status,
        })
    
    # Alerts and issues
    alerts = []
    
    # Check for branches with no contributions this month
    branches_no_contrib = branches.exclude(
        contributions__date__gte=month_start,
        contributions__date__lte=month_end
    ).count()
    
    if branches_no_contrib > 0:
        alerts.append({
            'type': 'warning',
            'message': f'{branches_no_contrib} branches have not submitted contributions this month'
        })
    
    # Check for overdue remittances
    overdue_remittances = Remittance.objects.filter(
        branch__in=branches,
        status='pending',
        year__lt=current_year
    ).count() + Remittance.objects.filter(
        branch__in=branches,
        status='pending',
        year=current_year,
        month__lt=current_month
    ).count()
    
    if overdue_remittances > 0:
        alerts.append({
            'type': 'danger',
            'message': f'{overdue_remittances} remittances are overdue'
        })
    
    # Check for unverified contributions
    unverified_contrib = Contribution.objects.filter(
        branch__in=branches,
        status='pending'
    ).count()
    
    if unverified_contrib > 0:
        alerts.append({
            'type': 'info',
            'message': f'{unverified_contrib} contributions pending verification'
        })
    
    # Calculate stats for display
    stats = {
        'total_contributions': month_contributions.get('total') or Decimal('0.00'),
        'contribution_count': month_contributions.get('count') or 0,
        'total_mission': month_contributions.get('total', Decimal('0.00')) * Decimal('0.10') if month_contributions.get('total') else Decimal('0.00'),
        'mission_percentage': 10,
        'total_expenditures': month_expenditures.get('total') or Decimal('0.00'),
        'expenditure_count': month_expenditures.get('count') or 0,
        'net_balance': (month_contributions.get('total') or Decimal('0.00')) - (month_expenditures.get('total') or Decimal('0.00')),
        'branch_count': total_branches,
    }
    
    context = {
        'total_branches': total_branches,
        'month_contributions': month_contributions,
        'month_expenditures': month_expenditures,
        'pending_remittances': pending_remittances,
        'closed_months': closed_months,
        'top_branches': top_branches,
        'recent_contributions': recent_contributions,
        'recent_expenditures': recent_expenditures,
        'branch_performance': branch_performance,
        'alerts': alerts,
        'stats': stats,
        'current_month': current_month,
        'current_year': current_year,
        'month_name': calendar.month_name[current_month],
        'areas': Area.objects.filter(is_active=True),
        'districts': District.objects.filter(is_active=True),
        'branches': Branch.objects.filter(is_active=True),
        'selected_area': area_id,
        'selected_district': district_id,
        'selected_branch': branch_id,
    }
    
    return render(request, 'auditing/auditor_dashboard.html', context)


@login_required
def financial_audit_report(request):
    """Comprehensive financial audit report."""
    if not (request.user.is_auditor or request.user.is_mission_admin):
        messages.error(request, 'Access denied.')
        return redirect('core:dashboard')
    
    # Get filters
    today = date.today()
    year = int(request.GET.get('year', today.year))
    month = request.GET.get('month')
    branch_id = request.GET.get('branch')
    area_id = request.GET.get('area')
    district_id = request.GET.get('district')
    
    # Date range
    if month:
        month = int(month)
        start_date = date(year, month, 1)
        if month == 12:
            end_date = date(year + 1, 1, 1) - timedelta(days=1)
        else:
            end_date = date(year, month + 1, 1) - timedelta(days=1)
        period_type = 'monthly'
    else:
        start_date = date(year, 1, 1)
        end_date = date(year, 12, 31)
        period_type = 'yearly'
    
    # DEPRECATED: Year-as-state architecture - Use date filtering only
    fiscal_year = None  # Not used anymore
    
    # Filter branches
    branches = Branch.objects.filter(is_active=True)
    if area_id:
        branches = branches.filter(district__area_id=area_id)
    if district_id:
        branches = branches.filter(district_id=district_id)
    if branch_id:
        branches = branches.filter(id=branch_id)
    
    # Collect data for each branch
    branch_data = []
    grand_totals = {
        'contributions': Decimal('0'),
        'mission_allocation': Decimal('0'),
        'branch_retained': Decimal('0'),
        'expenditures': Decimal('0'),
        'remittances_due': Decimal('0'),
        'remittances_sent': Decimal('0'),
        'payroll': Decimal('0'),
        'net_balance': Decimal('0'),
    }
    
    for branch in branches:
        # Contributions
        # DEPRECATED: Year-as-state architecture - Use date filtering only
        contributions = Contribution.objects.filter(
            branch=branch,
            date__gte=start_date,
            date__lte=end_date
        )
        
        contrib_total = contributions.aggregate(total=Sum('amount'))['total'] or Decimal('0')
        mission_alloc = contributions.aggregate(total=Sum('mission_amount'))['total'] or Decimal('0')
        branch_retained = contributions.aggregate(total=Sum('branch_amount'))['total'] or Decimal('0')
        
        # Expenditures
        # DEPRECATED: Year-as-state architecture - Use date filtering only
        expenditures = Expenditure.objects.filter(
            branch=branch,
            date__gte=start_date,
            date__lte=end_date
        )
        
        expend_total = expenditures.aggregate(total=Sum('amount'))['total'] or Decimal('0')
        
        # Remittances
        remittances = Remittance.objects.filter(branch=branch)
        if period_type == 'yearly':
            remittances = remittances.filter(year=year)
        else:
            remittances = remittances.filter(year=year, month=month)
        
        remit_due = remittances.aggregate(total=Sum('amount_due'))['total'] or Decimal('0')
        remit_sent = remittances.aggregate(total=Sum('amount_sent'))['total'] or Decimal('0')
        
        # Payroll
        payroll_runs = PayrollRun.objects.filter(year=year)
        if month:
            payroll_runs = payroll_runs.filter(month=month)
        
        payroll_total = PaySlip.objects.filter(
            payroll_run__in=payroll_runs,
            staff__user__branch=branch
        ).aggregate(total=Sum('net_pay'))['total'] or Decimal('0')
        
        # Calculate net balance
        net_balance = branch_retained - expend_total - payroll_total
        
        # Variance analysis
        variance = remit_sent - remit_due if remit_due > 0 else Decimal('0')
        variance_pct = (variance / remit_due * 100) if remit_due > 0 else Decimal('0')
        
        branch_data.append({
            'branch': branch,
            'contributions': contrib_total,
            'mission_allocation': mission_alloc,
            'branch_retained': branch_retained,
            'expenditures': expend_total,
            'remittances_due': remit_due,
            'remittances_sent': remit_sent,
            'payroll': payroll_total,
            'net_balance': net_balance,
            'variance': variance,
            'variance_pct': variance_pct,
            'compliance': 'Good' if variance >= 0 else 'Deficit',
        })
        
        # Update grand totals
        grand_totals['contributions'] += contrib_total
        grand_totals['mission_allocation'] += mission_alloc
        grand_totals['branch_retained'] += branch_retained
        grand_totals['expenditures'] += expend_total
        grand_totals['remittances_due'] += remit_due
        grand_totals['remittances_sent'] += remit_sent
        grand_totals['payroll'] += payroll_total
        grand_totals['net_balance'] += net_balance
    
    # Sort by contributions (highest first)
    branch_data.sort(key=lambda x: x['contributions'], reverse=True)
    
    context = {
        'branch_data': branch_data,
        'grand_totals': grand_totals,
        'period_type': period_type,
        'year': year,
        'month': month,
        'month_name': calendar.month_name[month] if month else None,
        'start_date': start_date,
        'end_date': end_date,
        'branches': Branch.objects.filter(is_active=True),
        'areas': Area.objects.filter(is_active=True),
        'districts': District.objects.filter(is_active=True),
        'selected_branch': branch_id,
        'selected_area': area_id,
        'selected_district': district_id,
        'available_years': list(range(today.year - 3, today.year + 2)),
        'months': [(i, calendar.month_name[i]) for i in range(1, 13)],
    }
    
    return render(request, 'auditing/financial_audit_report.html', context)


@login_required
def export_audit_report_excel(request):
    """Export audit report to Excel."""
    if not (request.user.is_auditor or request.user.is_mission_admin):
        messages.error(request, 'Access denied.')
        return redirect('core:dashboard')
    
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        messages.error(request, 'Excel export not available. Install openpyxl.')
        return redirect('auditing:financial_audit_report')
    
    # Get same data as report
    today = date.today()
    year = int(request.GET.get('year', today.year))
    month = request.GET.get('month')
    
    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Financial Audit Report"
    
    # Headers
    ws['A1'] = "SDSCC Financial Audit Report"
    ws['A1'].font = Font(size=16, bold=True)
    ws['A2'] = f"Period: {calendar.month_name[int(month)] if month else 'Full Year'} {year}"
    ws['A3'] = f"Generated: {today.strftime('%Y-%m-%d %H:%M')}"
    
    # Column headers
    headers = ['Branch', 'Contributions', 'Mission Allocation', 'Branch Retained', 
               'Expenditures', 'Remittances Due', 'Remittances Sent', 'Payroll', 'Net Balance']
    
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=5, column=col)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    
    # Add data (simplified - you'd get actual data here)
    row = 6
    # ... add branch data rows ...
    
    # Save to response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="audit_report_{year}_{month or "full"}.xlsx"'
    wb.save(response)
    
    return response


@login_required
def contribution_audit_trail(request):
    """Detailed audit trail for contributions."""
    if not (request.user.is_auditor or request.user.is_mission_admin):
        messages.error(request, 'Access denied.')
        return redirect('core:dashboard')
    
    # Get filters
    branch_id = request.GET.get('branch')
    contrib_type_id = request.GET.get('type')
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    
    contributions = Contribution.objects.select_related(
        'branch', 'contribution_type', 'member', 'created_by', 'verified_by'
    ).order_by('-date', '-created_at')
    
    if branch_id:
        contributions = contributions.filter(branch_id=branch_id)
    if contrib_type_id:
        contributions = contributions.filter(contribution_type_id=contrib_type_id)
    if start_date:
        contributions = contributions.filter(date__gte=start_date)
    if end_date:
        contributions = contributions.filter(date__lte=end_date)
    
    # Pagination
    from django.core.paginator import Paginator
    paginator = Paginator(contributions, 50)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'page_obj': page_obj,
        'branches': Branch.objects.filter(is_active=True),
        'contribution_types': ContributionType.objects.filter(is_active=True),
        'selected_branch': branch_id,
        'selected_type': contrib_type_id,
        'start_date': start_date,
        'end_date': end_date,
    }
    
    return render(request, 'auditing/contribution_audit_trail.html', context)


@login_required
def expenditure_audit_trail(request):
    """Detailed audit trail for expenditures."""
    if not (request.user.is_auditor or request.user.is_mission_admin):
        messages.error(request, 'Access denied.')
        return redirect('core:dashboard')
    
    # Get filters
    branch_id = request.GET.get('branch')
    category_id = request.GET.get('category')
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    status = request.GET.get('status')
    
    expenditures = Expenditure.objects.select_related(
        'branch', 'category', 'created_by', 'approved_by'
    ).order_by('-date', '-created_at')
    
    if branch_id:
        expenditures = expenditures.filter(branch_id=branch_id)
    if category_id:
        expenditures = expenditures.filter(category_id=category_id)
    if start_date:
        expenditures = expenditures.filter(date__gte=start_date)
    if end_date:
        expenditures = expenditures.filter(date__lte=end_date)
    if status:
        expenditures = expenditures.filter(status=status)
    
    # Pagination
    from django.core.paginator import Paginator
    paginator = Paginator(expenditures, 50)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'page_obj': page_obj,
        'branches': Branch.objects.filter(is_active=True),
        'categories': ExpenditureCategory.objects.filter(is_active=True),
        'selected_branch': branch_id,
        'selected_category': category_id,
        'start_date': start_date,
        'end_date': end_date,
        'selected_status': status,
    }
    
    return render(request, 'auditing/expenditure_audit_trail.html', context)


@login_required
def variance_analysis(request):
    """Variance analysis between budgets and actuals."""
    if not (request.user.is_auditor or request.user.is_mission_admin):
        messages.error(request, 'Access denied.')
        return redirect('core:dashboard')
    
    today = date.today()
    year = int(request.GET.get('year', today.year))
    
    # Get all branches
    branches = Branch.objects.filter(is_active=True)
    
    analysis_data = []
    
    for branch in branches:
        # Get monthly closes for the year
        monthly_closes = MonthlyClose.objects.filter(
            branch=branch,
            year=year,
            is_closed=True
        )
        
        for mc in monthly_closes:
            variance = mc.total_tithe - mc.target_amount if mc.target_amount > 0 else Decimal('0')
            variance_pct = (variance / mc.target_amount * 100) if mc.target_amount > 0 else Decimal('0')
            
            analysis_data.append({
                'branch': branch,
                'month': mc.month,
                'month_name': calendar.month_name[mc.month],
                'target': mc.target_amount,
                'actual': mc.total_tithe,
                'variance': variance,
                'variance_pct': variance_pct,
                'status': 'Exceeded' if variance > 0 else 'Below' if variance < 0 else 'Met',
            })
    
    context = {
        'analysis_data': analysis_data,
        'year': year,
        'available_years': list(range(today.year - 3, today.year + 2)),
    }
    
    return render(request, 'auditing/variance_analysis.html', context)
