"""
Member Export Views - Excel export functionality
"""

from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import HttpResponse, JsonResponse
from django.db.models import Q
from datetime import date

from accounts.models import User, UserProfile
from core.models import Branch, District, Area


@login_required
def export_members_page(request):
    """Page for member export options."""
    if not (request.user.is_any_admin or request.user.is_branch_executive):
        messages.error(request, 'Access denied.')
        return redirect('core:dashboard')
    
    # Get branches based on user role
    if request.user.is_mission_admin:
        branches = Branch.objects.filter(is_active=True)
        areas = Area.objects.filter(is_active=True)
        districts = District.objects.filter(is_active=True)
    elif request.user.is_area_executive and request.user.managed_area:
        branches = Branch.objects.filter(district__area=request.user.managed_area, is_active=True)
        areas = Area.objects.filter(id=request.user.managed_area.id)
        districts = District.objects.filter(area=request.user.managed_area, is_active=True)
    elif request.user.is_district_executive and request.user.managed_district:
        branches = Branch.objects.filter(district=request.user.managed_district, is_active=True)
        areas = Area.objects.none()
        districts = District.objects.filter(id=request.user.managed_district.id)
    else:
        branches = Branch.objects.filter(id=request.user.branch.id) if request.user.branch else Branch.objects.none()
        areas = Area.objects.none()
        districts = District.objects.none()
    
    context = {
        'branches': branches,
        'areas': areas,
        'districts': districts,
    }
    
    return render(request, 'members/export_page.html', context)


@login_required
def export_members_excel(request):
    """Export members to Excel file."""
    if not (request.user.is_any_admin or request.user.is_branch_executive):
        messages.error(request, 'Access denied.')
        return redirect('core:dashboard')
    
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        messages.error(request, 'Excel export not available. Install openpyxl: pip install openpyxl')
        return redirect('members:export_page')
    
    # Get filters
    branch_id = request.GET.get('branch')
    area_id = request.GET.get('area')
    district_id = request.GET.get('district')
    role = request.GET.get('role')
    status = request.GET.get('status', 'active')
    
    # Build query
    members = User.objects.select_related('branch', 'branch__district', 'branch__district__area')
    
    # Apply filters
    if status == 'active':
        members = members.filter(is_active=True)
    elif status == 'inactive':
        members = members.filter(is_active=False)
    
    if branch_id:
        members = members.filter(branch_id=branch_id)
    elif district_id:
        members = members.filter(branch__district_id=district_id)
    elif area_id:
        members = members.filter(branch__district__area_id=area_id)
    
    if role:
        members = members.filter(role=role)
    
    # Check permissions
    if not request.user.is_mission_admin:
        if request.user.is_area_executive and request.user.managed_area:
            members = members.filter(branch__district__area=request.user.managed_area)
        elif request.user.is_district_executive and request.user.managed_district:
            members = members.filter(branch__district=request.user.managed_district)
        elif request.user.branch:
            members = members.filter(branch=request.user.branch)
    
    members = members.order_by('branch__name', 'last_name', 'first_name')
    
    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Members"
    
    # Styles
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Title
    ws.merge_cells('A1:L1')
    title_cell = ws['A1']
    title_cell.value = "SDSCC Member Directory"
    title_cell.font = Font(bold=True, size=16)
    title_cell.alignment = Alignment(horizontal="center")
    
    # Subtitle
    ws.merge_cells('A2:L2')
    subtitle_cell = ws['A2']
    filter_text = []
    if branch_id:
        branch = Branch.objects.get(pk=branch_id)
        filter_text.append(f"Branch: {branch.name}")
    if district_id:
        district = District.objects.get(pk=district_id)
        filter_text.append(f"District: {district.name}")
    if area_id:
        area = Area.objects.get(pk=area_id)
        filter_text.append(f"Area: {area.name}")
    
    subtitle_cell.value = f"Exported: {date.today().strftime('%Y-%m-%d')} | {' | '.join(filter_text) if filter_text else 'All Members'}"
    subtitle_cell.alignment = Alignment(horizontal="center")
    
    # Headers
    headers = [
        'Member ID',
        'First Name',
        'Last Name',
        'Email',
        'Phone',
        'Gender',
        'Date of Birth',
        'Branch',
        'District',
        'Area',
        'Role',
        'Status'
    ]
    
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    
    # Data rows
    row = 5
    for member in members:
        ws.cell(row=row, column=1, value=member.member_id or member.user_id)
        ws.cell(row=row, column=2, value=member.first_name)
        ws.cell(row=row, column=3, value=member.last_name)
        ws.cell(row=row, column=4, value=member.email or '')
        ws.cell(row=row, column=5, value=member.phone or '')
        ws.cell(row=row, column=6, value=member.get_gender_display() if member.gender else '')
        ws.cell(row=row, column=7, value=member.date_of_birth.strftime('%Y-%m-%d') if member.date_of_birth else '')
        ws.cell(row=row, column=8, value=member.branch.name if member.branch else '')
        ws.cell(row=row, column=9, value=member.branch.district.name if member.branch and member.branch.district else '')
        ws.cell(row=row, column=10, value=member.branch.district.area.name if member.branch and member.branch.district and member.branch.district.area else '')
        ws.cell(row=row, column=11, value=member.get_role_display())
        ws.cell(row=row, column=12, value='Active' if member.is_active else 'Inactive')
        
        # Apply borders
        for col in range(1, 13):
            ws.cell(row=row, column=col).border = border
        
        row += 1
    
    # Adjust column widths
    column_widths = {
        'A': 15,  # Member ID
        'B': 15,  # First Name
        'C': 15,  # Last Name
        'D': 25,  # Email
        'E': 15,  # Phone
        'F': 10,  # Gender
        'G': 15,  # DOB
        'H': 20,  # Branch
        'I': 20,  # District
        'J': 20,  # Area
        'K': 15,  # Role
        'L': 10,  # Status
    }
    
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
    
    # Freeze header rows
    ws.freeze_panes = 'A5'
    
    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    
    filename = f"members_export_{date.today().strftime('%Y%m%d')}"
    if branch_id:
        branch = Branch.objects.get(pk=branch_id)
        filename += f"_{branch.code}"
    filename += ".xlsx"
    
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    wb.save(response)
    
    return response


@login_required
def export_members_preview(request):
    """Preview members export data before download."""
    try:
        # Check permissions
        if not (request.user.is_any_admin or request.user.is_branch_executive):
            return JsonResponse({
                'success': False,
                'message': 'Access denied.'
            }, status=403)

        # Get filters
        branch_id = request.GET.get('branch')
        area_id = request.GET.get('area')
        district_id = request.GET.get('district')
        role = request.GET.get('role')
        status_values = request.GET.getlist('status')
        fields = request.GET.getlist('fields')
        sort_by = request.GET.get('sort_by', 'name')

        # Default to active if no status specified
        if not status_values:
            status_values = ['active']

        # Build query with related data
        members = User.objects.select_related(
            'branch', 'branch__district', 'branch__district__area', 'profile'
        )

        # Apply status filters
        status_filters = Q()
        if 'active' in status_values:
            status_filters |= Q(is_active=True)
        if 'inactive' in status_values:
            status_filters |= Q(is_active=False)
        # Note: transferred and deceased might be handled differently in your system

        if status_filters:
            members = members.filter(status_filters)

        # Apply location filters
        if branch_id:
            members = members.filter(branch_id=branch_id)
        elif district_id:
            members = members.filter(branch__district_id=district_id)
        elif area_id:
            members = members.filter(branch__district__area_id=area_id)

        # Apply role filter
        if role:
            members = members.filter(role=role)

        # Apply permission-based filtering
        if not request.user.is_mission_admin:
            if request.user.is_area_executive and request.user.managed_area:
                members = members.filter(branch__district__area=request.user.managed_area)
            elif request.user.is_district_executive and request.user.managed_district:
                members = members.filter(branch__district=request.user.managed_district)
            elif request.user.branch:
                members = members.filter(branch=request.user.branch)

        # Apply sorting
        if sort_by == 'name':
            members = members.order_by('last_name', 'first_name')
        elif sort_by == 'date_joined':
            members = members.order_by('created_at')
        elif sort_by == 'branch':
            members = members.order_by('branch__name', 'last_name', 'first_name')
        elif sort_by == 'role':
            members = members.order_by('role', 'last_name', 'first_name')
        else:
            members = members.order_by('last_name', 'first_name')

        # Get total count
        total_count = members.count()

        # Build headers and sample data based on selected fields
        headers = []
        sample_rows = []

        # Default fields if none specified
        if not fields:
            fields = ['personal', 'contact', 'church']

        # Always include basic info
        if 'personal' in fields or 'contact' in fields:
            headers.extend(['Member ID', 'First Name', 'Last Name'])

        # Add fields based on selection
        if 'contact' in fields:
            headers.extend(['Email', 'Phone'])

        if 'address' in fields:
            headers.extend(['Address', 'City', 'Region'])

        if 'personal' in fields:
            headers.extend(['Gender', 'Date of Birth', 'Marital Status'])

        if 'church' in fields:
            headers.extend(['Branch', 'District', 'Area', 'Role', 'Status'])

        if 'occupation' in fields:
            headers.extend(['Profession', 'Employer'])

        if 'family' in fields:
            headers.extend(['Emergency Contact', 'Relationship', 'Emergency Phone'])

        if 'skills' in fields:
            headers.extend(['Skills', 'Talents'])

        if 'photo' in fields:
            headers.append('Photo URL')

        # Get sample data (first 5 members)
        sample_members = members[:5]

        for member in sample_members:
            row = []

            # Basic info
            if 'personal' in fields or 'contact' in fields:
                row.extend([
                    member.member_id or member.user_id,
                    member.first_name,
                    member.last_name
                ])

            # Contact info
            if 'contact' in fields:
                row.extend([
                    member.email or '',
                    member.phone or ''
                ])

            # Address info
            if 'address' in fields:
                profile = member.profile if hasattr(member, 'profile') else None
                row.extend([
                    profile.address if profile else '',
                    profile.city if profile else '',
                    profile.region if profile else ''
                ])

            # Personal info
            if 'personal' in fields:
                profile = member.profile if hasattr(member, 'profile') else None
                row.extend([
                    member.get_gender_display() if member.gender else '',
                    member.date_of_birth.strftime('%Y-%m-%d') if member.date_of_birth else '',
                    profile.get_marital_status_display() if profile and profile.marital_status else ''
                ])

            # Church info
            if 'church' in fields:
                row.extend([
                    member.branch.name if member.branch else '',
                    member.branch.district.name if member.branch and member.branch.district else '',
                    member.branch.district.area.name if member.branch and member.branch.district and member.branch.district.area else '',
                    member.get_role_display(),
                    'Active' if member.is_active else 'Inactive'
                ])

            # Occupation info
            if 'occupation' in fields:
                profile = member.profile if hasattr(member, 'profile') else None
                row.extend([
                    profile.profession if profile else '',
                    profile.employer if profile else ''
                ])

            # Family info
            if 'family' in fields:
                profile = member.profile if hasattr(member, 'profile') else None
                row.extend([
                    profile.emergency_contact_name if profile else '',
                    profile.emergency_contact_relationship if profile else '',
                    profile.emergency_contact_phone if profile else ''
                ])

            # Skills info
            if 'skills' in fields:
                profile = member.profile if hasattr(member, 'profile') else None
                row.extend([
                    profile.skills if profile else '',
                    profile.talents if profile else ''
                ])

            # Photo URL
            if 'photo' in fields:
                row.append(member.profile_picture.url if member.profile_picture else '')

            sample_rows.append(row)

        return JsonResponse({
            'success': True,
            'count': total_count,
            'headers': headers,
            'rows': sample_rows
        })

    except Exception as e:
        import traceback
        return JsonResponse({
            'success': False,
            'message': f'Error: {str(e)}',
            'traceback': traceback.format_exc()
        }, status=500)


@login_required
def export_members_csv(request):
    """Export members to CSV file."""
    if not (request.user.is_any_admin or request.user.is_branch_executive):
        messages.error(request, 'Access denied.')
        return redirect('core:dashboard')
    
    import csv
    
    # Get filters (same as Excel export)
    branch_id = request.GET.get('branch')
    area_id = request.GET.get('area')
    district_id = request.GET.get('district')
    role = request.GET.get('role')
    status = request.GET.get('status', 'active')
    
    # Build query
    members = User.objects.select_related('branch', 'branch__district', 'branch__district__area')
    
    # Apply filters
    if status == 'active':
        members = members.filter(is_active=True)
    elif status == 'inactive':
        members = members.filter(is_active=False)
    
    if branch_id:
        members = members.filter(branch_id=branch_id)
    elif district_id:
        members = members.filter(branch__district_id=district_id)
    elif area_id:
        members = members.filter(branch__district__area_id=area_id)
    
    if role:
        members = members.filter(role=role)
    
    # Check permissions
    if not request.user.is_mission_admin:
        if request.user.is_area_executive and request.user.managed_area:
            members = members.filter(branch__district__area=request.user.managed_area)
        elif request.user.is_district_executive and request.user.managed_district:
            members = members.filter(branch__district=request.user.managed_district)
        elif request.user.branch:
            members = members.filter(branch=request.user.branch)
    
    members = members.order_by('branch__name', 'last_name', 'first_name')
    
    # Create CSV response
    response = HttpResponse(content_type='text/csv')
    filename = f"members_export_{date.today().strftime('%Y%m%d')}"
    if branch_id:
        branch = Branch.objects.get(pk=branch_id)
        filename += f"_{branch.code}"
    filename += ".csv"
    
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    writer = csv.writer(response)
    
    # Write headers
    writer.writerow([
        'Member ID',
        'First Name',
        'Last Name',
        'Email',
        'Phone',
        'Gender',
        'Date of Birth',
        'Branch',
        'District',
        'Area',
        'Role',
        'Status'
    ])
    
    # Write data
    for member in members:
        writer.writerow([
            member.member_id or member.user_id,
            member.first_name,
            member.last_name,
            member.email or '',
            member.phone or '',
            member.get_gender_display() if member.gender else '',
            member.date_of_birth.strftime('%Y-%m-%d') if member.date_of_birth else '',
            member.branch.name if member.branch else '',
            member.branch.district.name if member.branch and member.branch.district else '',
            member.branch.district.area.name if member.branch and member.branch.district and member.branch.district.area else '',
            member.get_role_display(),
            'Active' if member.is_active else 'Inactive'
        ])
    
    return response
