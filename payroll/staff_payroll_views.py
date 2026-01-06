"""
Comprehensive Staff & Payroll Management Views
Handles staff management, salary configuration, and payroll processing
"""

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db.models import Sum, Q, Count
from django.http import JsonResponse
from django.utils import timezone
from datetime import date
from decimal import Decimal
import calendar

from accounts.models import User
from .models import StaffPayrollProfile, PayrollRun, PaySlip
from core.models import FiscalYear
from expenditure.models import Expenditure, ExpenditureCategory


@login_required
def staff_payroll_management(request):
    """
    Comprehensive Staff & Payroll Management Page
    Shows all staff with salary info, allows adding/editing staff, and managing payroll
    """
    if not request.user.is_mission_admin:
        messages.error(request, 'Access denied.')
        return redirect('core:dashboard')
    
    # Get all staff (pastors and staff role users)
    staff_profiles = StaffPayrollProfile.objects.select_related('user', 'user__branch').filter(is_active=True)
    
    # Get all users who are pastors or staff but don't have payroll profiles
    users_without_payroll = User.objects.filter(
        Q(role='pastor') | Q(role='staff'),
        is_active=True
    ).exclude(
        id__in=staff_profiles.values_list('user_id', flat=True)
    ).select_related('branch')
    
    # Calculate totals
    total_monthly_salary = staff_profiles.aggregate(total=Sum('base_salary'))['total'] or Decimal('0.00')
    total_staff_count = staff_profiles.count()
    
    # Get recent payroll runs
    recent_payrolls = PayrollRun.objects.all().order_by('-year', '-month')[:5]
    
    context = {
        'staff_with_payroll': staff_profiles,
        'users_without_payroll': users_without_payroll,
        'total_monthly_salary': total_monthly_salary,
        'total_staff_count': total_staff_count,
        'recent_payrolls': recent_payrolls,
    }
    
    return render(request, 'payroll/staff_payroll_management.html', context)


@login_required
def add_staff_to_payroll(request):
    """Add a user to payroll system."""
    from .models import StaffAllowance, AllowanceType
    from datetime import date
    
    if not request.user.is_mission_admin:
        messages.error(request, 'Access denied.')
        return redirect('core:dashboard')
    
    if request.method == 'POST':
        user_id = request.POST.get('user_id')
        base_salary = Decimal(request.POST.get('base_salary', '0'))
        tithe_deduction_percentage = Decimal(request.POST.get('tithe_deduction_percentage', '10'))
        housing_allowance = Decimal(request.POST.get('housing_allowance', '0'))
        transport_allowance = Decimal(request.POST.get('transport_allowance', '0'))
        other_allowances = Decimal(request.POST.get('other_allowances', '0'))
        position = request.POST.get('position', '')
        
        try:
            user = User.objects.get(pk=user_id)
            
            # Check if already exists
            if StaffPayrollProfile.objects.filter(user=user).exists():
                messages.warning(request, f'{user.get_full_name()} is already on payroll.')
                return redirect('payroll:staff_management')
            
            # Generate employee ID
            last_profile = StaffPayrollProfile.objects.order_by('-employee_id').first()
            if last_profile and last_profile.employee_id.startswith('EMP'):
                try:
                    last_num = int(last_profile.employee_id[3:])
                    employee_id = f'EMP{last_num + 1:04d}'
                except ValueError:
                    employee_id = f'EMP{StaffPayrollProfile.objects.count() + 1:04d}'
            else:
                employee_id = f'EMP{StaffPayrollProfile.objects.count() + 1:04d}'
            
            # Create payroll profile (without allowance fields)
            profile = StaffPayrollProfile.objects.create(
                user=user,
                employee_id=employee_id,
                base_salary=base_salary,
                tithe_deduction_percentage=tithe_deduction_percentage,
                position=position or user.get_role_display(),
                hire_date=date.today(),
                is_active=True
            )
            
            # Create allowances as separate records if provided
            today = date.today()
            
            if housing_allowance > 0:
                housing_type, _ = AllowanceType.objects.get_or_create(
                    code='HOUSING',
                    defaults={'name': 'Housing Allowance', 'is_taxable': False}
                )
                StaffAllowance.objects.create(
                    staff=profile,
                    allowance_type=housing_type,
                    amount=housing_allowance,
                    is_recurring=True,
                    start_date=today
                )
            
            if transport_allowance > 0:
                transport_type, _ = AllowanceType.objects.get_or_create(
                    code='TRANSPORT',
                    defaults={'name': 'Transport Allowance', 'is_taxable': False}
                )
                StaffAllowance.objects.create(
                    staff=profile,
                    allowance_type=transport_type,
                    amount=transport_allowance,
                    is_recurring=True,
                    start_date=today
                )
            
            if other_allowances > 0:
                other_type, _ = AllowanceType.objects.get_or_create(
                    code='OTHER',
                    defaults={'name': 'Other Allowances', 'is_taxable': False}
                )
                StaffAllowance.objects.create(
                    staff=profile,
                    allowance_type=other_type,
                    amount=other_allowances,
                    is_recurring=True,
                    start_date=today
                )
            
            messages.success(request, f'{user.get_full_name()} added to payroll successfully.')
        except User.DoesNotExist:
            messages.error(request, 'User not found.')
        except Exception as e:
            messages.error(request, f'Error adding to payroll: {str(e)}')
    
    return redirect('payroll:staff_management')


@login_required
def update_staff_salary(request, profile_id):
    """Update staff salary information."""
    from .models import StaffAllowance, AllowanceType
    from datetime import date
    
    if not request.user.is_mission_admin:
        messages.error(request, 'Access denied.')
        return redirect('core:dashboard')
    
    profile = get_object_or_404(StaffPayrollProfile, pk=profile_id)
    
    if request.method == 'POST':
        profile.base_salary = Decimal(request.POST.get('base_salary', profile.base_salary))
        profile.tithe_deduction_percentage = Decimal(request.POST.get('tithe_deduction_percentage', profile.tithe_deduction_percentage))
        profile.position = request.POST.get('position', profile.position)
        profile.save()
        
        # Update or create allowances
        housing_allowance = Decimal(request.POST.get('housing_allowance', '0'))
        transport_allowance = Decimal(request.POST.get('transport_allowance', '0'))
        other_allowances = Decimal(request.POST.get('other_allowances', '0'))
        today = date.today()
        
        # Housing allowance
        if housing_allowance > 0:
            housing_type, _ = AllowanceType.objects.get_or_create(
                code='HOUSING',
                defaults={'name': 'Housing Allowance', 'is_taxable': False}
            )
            allowance, created = StaffAllowance.objects.update_or_create(
                staff=profile,
                allowance_type=housing_type,
                defaults={'amount': housing_allowance, 'is_recurring': True, 'start_date': today}
            )
        else:
            StaffAllowance.objects.filter(staff=profile, allowance_type__code='HOUSING').delete()
        
        # Transport allowance
        if transport_allowance > 0:
            transport_type, _ = AllowanceType.objects.get_or_create(
                code='TRANSPORT',
                defaults={'name': 'Transport Allowance', 'is_taxable': False}
            )
            StaffAllowance.objects.update_or_create(
                staff=profile,
                allowance_type=transport_type,
                defaults={'amount': transport_allowance, 'is_recurring': True, 'start_date': today}
            )
        else:
            StaffAllowance.objects.filter(staff=profile, allowance_type__code='TRANSPORT').delete()
        
        # Other allowances
        if other_allowances > 0:
            other_type, _ = AllowanceType.objects.get_or_create(
                code='OTHER',
                defaults={'name': 'Other Allowances', 'is_taxable': False}
            )
            StaffAllowance.objects.update_or_create(
                staff=profile,
                allowance_type=other_type,
                defaults={'amount': other_allowances, 'is_recurring': True, 'start_date': today}
            )
        else:
            StaffAllowance.objects.filter(staff=profile, allowance_type__code='OTHER').delete()
        
        messages.success(request, f'Salary updated for {profile.user.get_full_name()}')
    
    return redirect('payroll:staff_management')


@login_required
def payroll_processing(request):
    """
    Payroll Processing Page
    Generate payroll for a month, mark as paid, view payslips
    """
    if not request.user.is_mission_admin:
        messages.error(request, 'Access denied.')
        return redirect('core:dashboard')
    
    if request.method == 'POST':
        action = request.POST.get('action')
        
        if action == 'generate':
            month = int(request.POST.get('month'))
            year = int(request.POST.get('year'))
            
            # Check if payroll already exists
            if PayrollRun.objects.filter(month=month, year=year).exists():
                messages.warning(request, f'Payroll for {calendar.month_name[month]} {year} already exists.')
                return redirect('payroll:payroll_processing')
            
            try:
                # DEPRECATED: Year-as-state architecture - fiscal_year no longer assigned
                # Payroll runs now use date-based filtering only
                
                # Create payroll run
                payroll_run = PayrollRun.objects.create(
                    month=month,
                    year=year,
                    status='draft',
                    processed_by=request.user,
                    processed_at=timezone.now()
                )
                
                # Generate payslips for all active staff
                from .models import StaffAllowance, StaffDeduction
                
                staff_profiles = StaffPayrollProfile.objects.filter(is_active=True).prefetch_related('allowances', 'deductions')
                payslip_count = 0
                total_net = Decimal('0.00')
                
                for profile in staff_profiles:
                    # Calculate total allowances from StaffAllowance records
                    allowances = profile.allowances.filter(is_recurring=True, end_date__isnull=True) | \
                                profile.allowances.filter(is_recurring=True, end_date__gte=date.today())
                    total_allowances = sum(a.amount for a in allowances)
                    
                    # Calculate total deductions from StaffDeduction records
                    deductions = profile.deductions.filter(is_recurring=True, end_date__isnull=True) | \
                                profile.deductions.filter(is_recurring=True, end_date__gte=date.today())
                    total_deductions = sum(d.amount for d in deductions)
                    
                    # Calculate tithe deduction
                    tithe_deduction = (profile.base_salary * profile.tithe_deduction_percentage) / Decimal('100')
                    
                    # Add tithe to total deductions
                    total_deductions += tithe_deduction
                    
                    # Calculate net pay
                    net_pay = profile.base_salary + total_allowances - total_deductions
                    
                    # Create payslip
                    payslip = PaySlip.objects.create(
                        staff=profile,
                        payroll_run=payroll_run,
                        base_salary=profile.base_salary,
                        allowances=total_allowances,
                        deductions=total_deductions,
                        net_pay=net_pay,
                        status='draft'
                    )
                    
                    # Post tithe deduction to mission ledger
                    if tithe_deduction > 0:
                        # Get or create tithe contribution type
                        tithe_type, _ = ContributionType.objects.get_or_create(
                            name='Staff Tithe',
                            defaults={
                                'name': 'Staff Tithe Deduction',
                                'category': 'tithe',
                                'scope': 'mission',
                                'mission_percentage': 100,
                                'area_percentage': 0,
                                'district_percentage': 0,
                                'branch_percentage': 0,
                                'is_individual': True,
                                'is_general': False,
                                'frequency': 'monthly'
                            }
                        )
                        
                        # Create contribution record for the tithe
                        Contribution.objects.create(
                            contribution_type=tithe_type,
                            branch=profile.user.branch,
                            member=profile.user,
                            date=date(year, month, 1),  # First day of the month
                            amount=tithe_deduction,
                            description=f'Tithe deduction from {profile.user.get_full_name()} salary for {calendar.month_name[month]} {year}',
                            reference=f'PAYROLL-{payroll_run.id}',
                            status='verified',
                            verified_by=request.user,
                            verified_at=timezone.now()
                        )
                
                # Update payroll run totals
                payroll_run.calculate_totals()
                payroll_run.save()
                
                messages.success(request, f'Generated {payslip_count} payslips for {calendar.month_name[month]} {year}')
            except Exception as e:
                messages.error(request, f'Error generating payroll: {str(e)}')
                return redirect('payroll:payroll_processing')
        
        elif action == 'change_status':
            payslip_id = request.POST.get('payslip_id')
            new_status = request.POST.get('new_status')
            
            if not payslip_id or not new_status:
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    return JsonResponse({'success': False, 'message': 'Payslip ID and new status are required.'})
                messages.error(request, 'Payslip ID and new status are required.')
                return redirect('payroll:payroll_processing')
            
            try:
                payslip = PaySlip.objects.get(id=payslip_id)
                old_status = payslip.status
                
                # Validate status change - allow skipping but prevent invalid transitions
                if new_status == 'cancelled' and payslip.status == 'paid':
                    error_msg = 'Cannot cancel a paid payslip.'
                    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                        return JsonResponse({'success': False, 'message': error_msg})
                    messages.error(request, error_msg)
                    return redirect('payroll:payroll_processing')
                
                # Allow status skipping (e.g., draft -> paid)
                # When skipping to paid, automatically mark as processed
                if new_status == 'paid':
                    if payslip.status != 'approved':
                        # Auto-approve if not already approved
                        payslip.status = 'approved'
                        payslip.save()
                        # Now open payment modal (handled in frontend)
                    
                    # Return payment modal request instead of marking as paid
                    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                        return JsonResponse({
                            'success': True, 
                            'message': f'Payslip for {payslip.staff.user.get_full_name()} is ready for payment',
                            'action': 'open_payment_modal',
                            'payslip_id': payslip.id,
                            'staff_name': payslip.staff.user.get_full_name(),
                            'amount': str(payslip.net_pay)
                        })
                    else:
                        messages.success(request, f'Payslip approved for {payslip.staff.user.get_full_name()}. Please complete payment.')
                        return redirect('payroll:payroll_processing')
                
                # For other status changes, allow direct updates
                payslip.status = new_status
                payslip.save()
                
                # Calculate updated status counts
                if payroll_run := payslip.payroll_run:
                    payslips = payroll_run.payslips.all()
                    status_counts = {
                        'draft': payslips.filter(status='draft').count(),
                        'processing': payslips.filter(status='processing').count(),
                        'approved': payslips.filter(status='approved').count(),
                        'paid': payslips.filter(status='paid').count()
                    }
                else:
                    status_counts = {}
                
                success_msg = f'Status changed for {payslip.staff.user.get_full_name()} from {old_status} to {new_status}'
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    return JsonResponse({
                        'success': True, 
                        'message': success_msg,
                        'status_counts': status_counts
                    })
                else:
                    messages.success(request, success_msg)
                    return redirect('payroll:payroll_processing')
                    
            except PaySlip.DoesNotExist:
                error_msg = 'Payslip not found.'
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    return JsonResponse({'success': False, 'message': error_msg})
                messages.error(request, error_msg)
            except Exception as e:
                error_msg = f'Error changing status: {str(e)}'
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    return JsonResponse({'success': False, 'message': error_msg})
                messages.error(request, error_msg)
            
            return redirect('payroll:payroll_processing')
        
        elif action == 'mark_processing':
            payslip_id = request.POST.get('payslip_id')
            payroll_id = request.POST.get('payroll_id')
            
            try:
                if payslip_id:
                    # Single payslip
                    payslip = PaySlip.objects.get(id=payslip_id)
                    if payslip.status != 'draft':
                        error_msg = 'Only draft payslips can be marked as processing.'
                        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                            return JsonResponse({'success': False, 'message': error_msg})
                        messages.error(request, error_msg)
                        return redirect('payroll:payroll_processing')
                    
                    payslip.status = 'processing'
                    payslip.save()
                    success_msg = f'Payslip marked as processing for {payslip.staff.user.get_full_name()}'
                    
                    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                        # Calculate updated status counts
                        if payroll_run := payslip.payroll_run:
                            payslips = payroll_run.payslips.all()
                            status_counts = {
                                'draft': payslips.filter(status='draft').count(),
                                'processing': payslips.filter(status='processing').count(),
                                'approved': payslips.filter(status='approved').count(),
                                'paid': payslips.filter(status='paid').count()
                            }
                        else:
                            status_counts = {}
                        
                        return JsonResponse({
                            'success': True, 
                            'message': success_msg,
                            'status_counts': status_counts
                        })
                    else:
                        messages.success(request, success_msg)
                        
                elif payroll_id:
                    # All payslips in payroll run
                    payroll_run = PayrollRun.objects.get(id=payroll_id)
                    draft_payslips = PaySlip.objects.filter(payroll_run=payroll_run, status='draft')
                    count = draft_payslips.count()
                    
                    if count > 0:
                        draft_payslips.update(status='processing')
                        success_msg = f'Marked {count} payslips as processing'
                    else:
                        success_msg = f'No draft payslips found to process'
                    
                    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                        # Calculate updated status counts
                        payslips = payroll_run.payslips.all()
                        status_counts = {
                            'draft': payslips.filter(status='draft').count(),
                            'processing': payslips.filter(status='processing').count(),
                            'approved': payslips.filter(status='approved').count(),
                            'paid': payslips.filter(status='paid').count()
                        }
                        
                        return JsonResponse({
                            'success': True, 
                            'message': success_msg,
                            'status_counts': status_counts
                        })
                    else:
                        messages.success(request, success_msg)
                        
            except PaySlip.DoesNotExist:
                error_msg = 'Payslip not found.'
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    return JsonResponse({'success': False, 'message': error_msg})
                messages.error(request, error_msg)
            except Exception as e:
                error_msg = f'Error updating status: {str(e)}'
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    return JsonResponse({'success': False, 'message': error_msg})
                messages.error(request, error_msg)
            
            if not request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return redirect('payroll:payroll_processing')
        
        elif action == 'mark_approved':
            payslip_id = request.POST.get('payslip_id')
            payroll_id = request.POST.get('payroll_id')
            
            try:
                if payslip_id:
                    # Single payslip
                    payslip = PaySlip.objects.get(id=payslip_id)
                    if payslip.status not in ['draft', 'processing']:
                        error_msg = 'Only draft or processing payslips can be approved.'
                        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                            return JsonResponse({'success': False, 'message': error_msg})
                        messages.error(request, error_msg)
                        return redirect('payroll:payroll_processing')
                    
                    payslip.status = 'approved'
                    payslip.save()
                    success_msg = f'Payslip approved for {payslip.staff.user.get_full_name()}'
                    
                    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                        # Calculate updated status counts
                        if payroll_run := payslip.payroll_run:
                            payslips = payroll_run.payslips.all()
                            status_counts = {
                                'draft': payslips.filter(status='draft').count(),
                                'processing': payslips.filter(status='processing').count(),
                                'approved': payslips.filter(status='approved').count(),
                                'paid': payslips.filter(status='paid').count()
                            }
                        else:
                            status_counts = {}
                        
                        return JsonResponse({
                            'success': True, 
                            'message': success_msg,
                            'status_counts': status_counts
                        })
                    else:
                        messages.success(request, success_msg)
                        
                elif payroll_id:
                    # All payslips in payroll run
                    payroll_run = PayrollRun.objects.get(id=payroll_id)
                    eligible_payslips = PaySlip.objects.filter(payroll_run=payroll_run, status__in=['draft', 'processing'])
                    count = eligible_payslips.count()
                    
                    if count > 0:
                        eligible_payslips.update(status='approved')
                        success_msg = f'Approved {count} payslips'
                    else:
                        success_msg = f'No draft or processing payslips found to approve'
                    
                    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                        # Calculate updated status counts
                        payslips = payroll_run.payslips.all()
                        status_counts = {
                            'draft': payslips.filter(status='draft').count(),
                            'processing': payslips.filter(status='processing').count(),
                            'approved': payslips.filter(status='approved').count(),
                            'paid': payslips.filter(status='paid').count()
                        }
                        
                        return JsonResponse({
                            'success': True, 
                            'message': success_msg,
                            'status_counts': status_counts
                        })
                    else:
                        messages.success(request, success_msg)
                        
            except PaySlip.DoesNotExist:
                error_msg = 'Payslip not found.'
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    return JsonResponse({'success': False, 'message': error_msg})
                messages.error(request, error_msg)
            except Exception as e:
                error_msg = f'Error approving payslips: {str(e)}'
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    return JsonResponse({'success': False, 'message': error_msg})
                messages.error(request, error_msg)
            
            if not request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return redirect('payroll:payroll_processing')
        
        elif action == 'mark_cancelled':
            payslip_id = request.POST.get('payslip_id')
            
            try:
                payslip = PaySlip.objects.get(id=payslip_id)
                if payslip.status == 'paid':
                    messages.error(request, 'Cannot cancel a paid payslip.')
                else:
                    payslip.status = 'cancelled'
                    payslip.save()
                    messages.success(request, f'Payslip cancelled for {payslip.staff.user.get_full_name()}')
            except Exception as e:
                messages.error(request, f'Error cancelling payslip: {str(e)}')
        
        elif action == 'mark_all_paid':
            payroll_id = request.POST.get('payroll_id')
            payment_date = date.today()
            payment_ref = f'PAY-{date.today().strftime("%Y%m%d")}-BATCH'
            
            try:
                payroll_run = PayrollRun.objects.get(id=payroll_id)
                payslips = PaySlip.objects.filter(payroll_run=payroll_run, status='approved')
                
                if not payslips.exists():
                    error_msg = 'No approved payslips found to mark as paid.'
                    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                        return JsonResponse({'success': False, 'message': error_msg})
                    messages.error(request, error_msg)
                    return redirect('payroll:payroll_processing')
                
                # Create expenditure category for payroll if not exists
                category, _ = ExpenditureCategory.objects.get_or_create(
                    name='Staff Salary',
                    defaults={
                        'code': 'SAL',
                        'category_type': 'payroll',
                        'scope': 'mission',
                        'is_active': True
                    }
                )
                
                # DEPRECATED: Year-as-state architecture - fiscal_year no longer assigned
                # Payroll expenditures now use date-based filtering only
                
                for payslip in payslips:
                    # Create expenditure record for audit trail
                    Expenditure.objects.create(
                        branch=payslip.staff.user.branch if payslip.staff.user.branch else None,
                        # fiscal_year=fiscal_year,  # REMOVED: Use date filtering instead
                        category=category,
                        level='mission',
                        amount=payslip.net_pay,
                        date=payment_date,
                        title=f'Salary - {payslip.staff.user.get_full_name()}',
                        description=f'Salary payment for {calendar.month_name[payslip.payroll_run.month]} {payslip.payroll_run.year}',
                        vendor=payslip.staff.user.get_full_name(),
                        reference_number=payment_ref,
                        status='approved',
                        approved_by=request.user,
                        approved_at=timezone.now(),
                        notes=f'Payment method: Bank transfer. Payslip ID: {payslip.id}'
                    )
                    
                    # Mark payslip as paid
                    payslip.status = 'paid'
                    payslip.payment_date = payment_date
                    payslip.payment_reference = payment_ref
                    payslip.save()
                
                success_msg = f'Marked {len(payslips)} payslips as paid.'
                
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    # Calculate updated status counts
                    all_payslips = payroll_run.payslips.all()
                    status_counts = {
                        'draft': all_payslips.filter(status='draft').count(),
                        'processing': all_payslips.filter(status='processing').count(),
                        'approved': all_payslips.filter(status='approved').count(),
                        'paid': all_payslips.filter(status='paid').count()
                    }
                    
                    return JsonResponse({
                        'success': True, 
                        'message': success_msg,
                        'status_counts': status_counts
                    })
                else:
                    messages.success(request, success_msg)
                    
            except PayrollRun.DoesNotExist:
                error_msg = 'Payroll run not found.'
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    return JsonResponse({'success': False, 'message': error_msg})
                messages.error(request, error_msg)
            except Exception as e:
                error_msg = f'Error marking payslips as paid: {str(e)}'
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    return JsonResponse({'success': False, 'message': error_msg})
                messages.error(request, error_msg)
            
            if not request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return redirect('payroll:payroll_processing')
        
        elif action == 'mark_paid':
            payslip_id = request.POST.get('payslip_id')
            payment_date = request.POST.get('payment_date')
            payment_reference = request.POST.get('payment_reference')
            payment_method = request.POST.get('payment_method')
            payment_notes = request.POST.get('payment_notes', '')
            
            if not payslip_id or not payment_date or not payment_reference:
                messages.error(request, 'Payment date and reference are required.')
                return redirect('payroll:payroll_processing')
            
            try:
                payslip = PaySlip.objects.get(id=payslip_id, status='pending')
                
                # Create expenditure category for payroll if not exists
                category, _ = ExpenditureCategory.objects.get_or_create(
                    name='Staff Salary',
                    defaults={
                        'code': 'SAL',
                        'category_type': 'payroll',
                        'scope': 'mission',
                        'is_active': True
                    }
                )
                
                # Create expenditure record for audit trail
                Expenditure.objects.create(
                    branch=payslip.staff.user.branch if payslip.staff.user.branch else None,
                    category=category,
                    level='mission',
                    amount=payslip.net_pay,
                    date=payment_date,
                    title=f'Salary - {payslip.staff.user.get_full_name()}',
                    description=f'Salary payment for {calendar.month_name[payslip.payroll_run.month]} {payslip.payroll_run.year}',
                    vendor=payslip.staff.user.get_full_name(),
                    reference_number=payment_reference,
                    status='approved',
                    approved_by=request.user,
                    approved_at=timezone.now(),
                    notes=f'Payment method: {payment_method}. Payslip ID: {payslip.id}. {payment_notes}'
                )
                
                # Mark payslip as paid
                payslip.status = 'paid'
                payslip.payment_date = payment_date
                payslip.payment_reference = payment_reference
                payslip.notes = payment_notes
                payslip.save()
                
                messages.success(request, f'Payment processed for {payslip.staff.user.get_full_name()}')
            except PaySlip.DoesNotExist:
                messages.error(request, 'Payslip not found or already paid.')
            except Exception as e:
                messages.error(request, f'Error processing payment: {str(e)}')
        
        return redirect('payroll:payroll_processing')
    
    # Get payroll runs - use timezone-aware date
    from django.utils import timezone
    now = timezone.now()
    month = int(request.GET.get('month', now.month))
    year = int(request.GET.get('year', now.year))
    
    payroll_run = PayrollRun.objects.filter(month=month, year=year).prefetch_related('payslips__staff__user').first()
    
    # Calculate status counts if payroll exists
    draft_count = processing_count = approved_count = paid_count = 0
    if payroll_run:
        payslips = payroll_run.payslips.all()
        draft_count = payslips.filter(status='draft').count()
        processing_count = payslips.filter(status='processing').count()
        approved_count = payslips.filter(status='approved').count()
        paid_count = payslips.filter(status='paid').count()
    
    # Month/year options
    months = [(i, calendar.month_name[i]) for i in range(1, 13)]
    years = list(range(now.year - 2, now.year + 3))
    years.sort(reverse=True)  # Sort years in descending order
    
    context = {
        'payroll_run': payroll_run,
        'draft_count': draft_count,
        'processing_count': processing_count,
        'approved_count': approved_count,
        'paid_count': paid_count,
        'month': month,
        'year': year,
        'month_name': calendar.month_name[month],
        'months': months,
        'years': years,
    }
    
    return render(request, 'payroll/payroll_processing.html', context)


@login_required
def my_payroll(request):
    """
    Pastor/Staff view of their own payroll
    Shows their salary, payslips, and payment history
    """
    if not (request.user.is_pastor or request.user.is_staff):
        messages.error(request, 'Access denied.')
        return redirect('core:dashboard')
    
    try:
        staff_profile = StaffPayrollProfile.objects.get(user=request.user, is_active=True)
    except StaffPayrollProfile.DoesNotExist:
        staff_profile = None
    
    # Get payslips
    payslips = PaySlip.objects.filter(
        staff__user=request.user
    ).select_related('payroll_run').order_by('-payroll_run__year', '-payroll_run__month')
    
    # Calculate totals
    total_paid = payslips.filter(status='paid').aggregate(total=Sum('net_pay'))['total'] or Decimal('0.00')
    total_pending = payslips.filter(status='pending').aggregate(total=Sum('net_pay'))['total'] or Decimal('0.00')
    
    # Get available years for filter
    available_years = list(set(payslip.payroll_run.year for payslip in payslips))
    available_years.sort(reverse=True)
    
    # Add month names to payslips
    import calendar
    for payslip in payslips:
        payslip.payroll_run.month_name = calendar.month_name[payslip.payroll_run.month]
    
    context = {
        'staff_profile': staff_profile,
        'payslips': payslips,
        'total_paid': total_paid,
        'total_pending': total_pending,
        'available_years': available_years,
    }
    
    return render(request, 'payroll/my_payroll.html', context)


@login_required
def payslip_detail(request, payslip_id):
    """View detailed payslip."""
    if not (request.user.is_pastor or request.user.is_staff or request.user.is_mission_admin):
        messages.error(request, 'Access denied.')
        return redirect('core:dashboard')
    
    # Admin can view any payslip, staff can only view their own
    if request.user.is_mission_admin:
        payslip = get_object_or_404(
            PaySlip.objects.select_related('staff', 'staff__user', 'payroll_run'),
            id=payslip_id
        )
    else:
        payslip = get_object_or_404(
            PaySlip.objects.select_related('staff', 'staff__user', 'payroll_run'),
            id=payslip_id,
            staff__user=request.user
        )
    
    # Add month name
    import calendar
    payslip.payroll_run.month_name = calendar.month_name[payslip.payroll_run.month]
    
    # Get site settings for church branding
    from core.models import SiteSettings
    site_settings = SiteSettings.objects.first()
    
    return render(request, 'payroll/payslip_detail.html', {
        'payslip': payslip,
        'site_settings': site_settings
    })


@login_required
def payment_history(request):
    """Payment history page showing monthly salary breakdown with filtering."""
    from decimal import Decimal
    from calendar import month_name
    
    # Get filter parameters
    year = request.GET.get('year', str(timezone.now().year))
    month = request.GET.get('month', '')
    staff_id = request.GET.get('staff')
    
    # Check permissions
    is_admin = request.user.is_mission_admin
    is_auditor = request.user.is_auditor and SiteSettings.get_settings().allow_auditor_payroll_access
    is_own_record = False
    
    if not (is_admin or is_auditor):
        # Regular staff can only view their own records
        try:
            profile = StaffPayrollProfile.objects.get(user=request.user, is_active=True)
            is_own_record = True
            staff_id = str(profile.id)
        except StaffPayrollProfile.DoesNotExist:
            messages.error(request, 'You do not have a payroll profile.')
            return redirect('core:dashboard')
    
    # Base queryset
    payslips = PaySlip.objects.select_related(
        'staff__user',
        'staff',
        'payroll_run'
    ).order_by('-payroll_run__year', '-payroll_run__month', 'staff__user__last_name')
    
    # Apply filters
    if year:
        payslips = payslips.filter(payroll_run__year=year)
    
    if month:
        payslips = payslips.filter(payroll_run__month=month)
    
    if staff_id:
        payslips = payslips.filter(staff_id=staff_id)
    
    # Calculate summary statistics
    total_gross_pay = payslips.aggregate(total=Sum('gross_pay'))['total'] or Decimal('0.00')
    total_deductions = payslips.aggregate(total=Sum('total_deductions'))['total'] or Decimal('0.00')
    total_net_pay = payslips.aggregate(total=Sum('net_pay'))['total'] or Decimal('0.00')
    total_payslips = payslips.count()
    
    # Get monthly breakdown
    monthly_breakdown = payslips.values(
        'payroll_run__year',
        'payroll_run__month'
    ).annotate(
        total_gross=Sum('gross_pay'),
        total_deductions=Sum('total_deductions'),
        total_net=Sum('net_pay'),
        payslip_count=Count('id')
    ).order_by('-payroll_run__year', '-payroll_run__month')
    
    # Get staff list (for filter)
    if is_admin or is_auditor:
        staff_profiles = StaffPayrollProfile.objects.filter(is_active=True).select_related('user')
    else:
        staff_profiles = StaffPayrollProfile.objects.filter(user=request.user, is_active=True)
    
    # Pagination
    from django.core.paginator import Paginator
    paginator = Paginator(payslips, 50)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'payslips': page_obj,
        'staff_profiles': staff_profiles,
        'selected_year': year,
        'selected_month': month,
        'selected_staff': staff_id,
        'years': range(timezone.now().year - 3, timezone.now().year + 1),
        'months': [(i, month_name[i]) for i in range(1, 13)],
        
        # Summary
        'total_gross_pay': total_gross_pay,
        'total_deductions': total_deductions,
        'total_net_pay': total_net_pay,
        'total_payslips': total_payslips,
        
        # Monthly breakdown
        'monthly_breakdown': monthly_breakdown,
        
        # Permissions
        'is_admin': is_admin,
        'is_auditor': is_auditor,
        'is_own_record': is_own_record,
    }
    
    return render(request, 'payroll/payment_history.html', context)


@login_required
def export_payment_history(request):
    """Export payment history to Excel."""
    from decimal import Decimal
    from calendar import month_name
    
    # Get filter parameters
    year = request.GET.get('year', str(timezone.now().year))
    month = request.GET.get('month', '')
    staff_id = request.GET.get('staff')
    
    # Check permissions
    is_admin = request.user.is_mission_admin
    is_auditor = request.user.is_auditor and SiteSettings.get_settings().allow_auditor_payroll_access
    
    if not (is_admin or is_auditor):
        # Regular staff can only export their own records
        try:
            profile = StaffPayrollProfile.objects.get(user=request.user, is_active=True)
            staff_id = str(profile.id)
        except StaffPayrollProfile.DoesNotExist:
            messages.error(request, 'You do not have a payroll profile.')
            return redirect('payroll:payment_history')
    
    # Base queryset
    payslips = PaySlip.objects.select_related(
        'staff__user',
        'staff',
        'payroll_run'
    ).order_by('-payroll_run__year', '-payroll_run__month', 'staff__user__last_name')
    
    # Apply filters
    if year:
        payslips = payslips.filter(payroll_run__year=year)
    if month:
        payslips = payslips.filter(payroll_run__month=month)
    if staff_id:
        payslips = payslips.filter(staff_id=staff_id)
    
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        messages.error(request, 'Excel export requires openpyxl package.')
        return redirect('payroll:payment_history')
    
    # Create workbook
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    
    # Payment History Sheet
    ws = wb.create_sheet("Payment History", 0)
    
    # Headers styling
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Headers
    headers = ["Year", "Month", "Staff ID", "Staff Name", "Position", 
               "Basic Salary", "Allowances", "Gross Pay", "Tithe", "Other Deductions", 
               "Total Deductions", "Net Pay", "Payment Status"]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Data rows
    for row, payslip in enumerate(payslips, 2):
        # Get tithe amount from deductions breakdown
        tithe_amount = 0
        if payslip.deductions_breakdown:
            for name, amount in payslip.deductions_breakdown.items():
                if 'Tithe' in name:
                    tithe_amount = Decimal(str(amount))
                    break
        
        data = [
            payslip.payroll_run.year,
            month_name[payslip.payroll_run.month],
            payslip.staff.user.username,
            payslip.staff.user.get_full_name(),
            payslip.staff.position or "N/A",
            payslip.basic_salary,
            payslip.total_allowances,
            payslip.gross_pay,
            tithe_amount,
            payslip.total_deductions - tithe_amount,
            payslip.total_deductions,
            payslip.net_pay,
            payslip.get_status_display()
        ]
        
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=value)
            if col >= 6 and col <= 12:  # Amount columns
                cell.number_format = '#,##0.00'
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"payment_history_{year}_{month or 'all'}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    wb.save(response)
    return response


@login_required
def export_payroll(request):
    """Export payroll to Excel."""
    if not request.user.is_mission_admin:
        messages.error(request, 'Access denied.')
        return redirect('core:dashboard')
    
    month = int(request.GET.get('month', date.today().month))
    year = int(request.GET.get('year', date.today().year))
    
    payroll_run = PayrollRun.objects.filter(month=month, year=year).prefetch_related('payslips__staff__user').first()
    
    if not payroll_run:
        messages.error(request, 'No payroll found for the selected period.')
        return redirect('payroll:payroll_processing')
    
    # Create Excel file
    import pandas as pd
    from django.http import HttpResponse
    
    # Get site settings for church branding
    from core.models import SiteSettings
    site_settings = SiteSettings.objects.first()
    church_name = site_settings.site_name if site_settings and site_settings.site_name != 'SDSCC' else (site_settings.tagline if site_settings else 'Seventh Day Sabbath Church of Christ')
    
    data = []
    for payslip in payroll_run.payslips.all():
        data.append({
            'Employee ID': payslip.staff.employee_id or '',
            'Employee Name': payslip.staff.user.get_full_name(),
            'Position': payslip.staff.position or '',
            'Email': payslip.staff.user.email,
            'Phone': payslip.staff.user.phone or '',
            'Base Salary': payslip.base_salary,
            'Total Allowances': payslip.total_allowances,
            'Gross Pay': payslip.gross_pay,
            'Total Deductions': payslip.total_deductions,
            'Net Pay': payslip.net_pay,
            'Status': payslip.get_status_display(),
            'Payment Date': payslip.payment_date or '',
            'Payment Reference': payslip.payment_reference or '',
        })
    
    df = pd.DataFrame(data)
    
    # Create response
    import calendar
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{church_name.replace(" ", "_")}_Payroll_{calendar.month_name[month]}_{year}.xlsx"'
    
    # Save to response
    with pd.ExcelWriter(response, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Payroll')
        
        # Add summary sheet
        summary_data = {
            'Metric': ['Church Name', 'Pay Period', 'Total Employees', 'Total Gross Pay', 'Total Deductions', 'Total Net Pay', 'Paid Employees', 'Pending Employees'],
            'Value': [
                church_name,
                f'{calendar.month_name[month]} {year}',
                payroll_run.payslips.count(),
                payroll_run.total_gross,
                payroll_run.total_deductions,
                payroll_run.total_net_pay,
                payroll_run.payslips.filter(status='paid').count(),
                payroll_run.payslips.filter(status='pending').count(),
            ]
        }
        summary_df = pd.DataFrame(summary_data)
        summary_df.to_excel(writer, index=False, sheet_name='Summary')
    
    return response


@login_required
def download_payslip(request, payslip_id):
    """Download payslip as PDF."""
    if not (request.user.is_pastor or request.user.is_staff):
        messages.error(request, 'Access denied.')
        return redirect('core:dashboard')
    
    payslip = get_object_or_404(
        PaySlip.objects.select_related('staff', 'staff__user', 'payroll_run'),
        id=payslip_id,
        staff__user=request.user
    )
    
    # For now, return a simple text response
    # In production, you would generate a PDF here
    response = HttpResponse(content_type='text/plain')
    response['Content-Disposition'] = f'attachment; filename="payslip_{payslip.staff.employee_id}_{payslip.payroll_run.month}_{payslip.payroll_run.year}.txt"'
    
    import calendar
    
    # Get site settings for church branding
    from core.models import SiteSettings
    site_settings = SiteSettings.objects.first()
    church_name = site_settings.site_name if site_settings and site_settings.site_name != 'SDSCC' else (site_settings.tagline if site_settings else 'Seventh Day Sabbath Church of Christ')
    
    content = f"""
{church_name}
OFFICIAL PAYSLIP - {payslip.payroll_run.month_name} {payslip.payroll_run.year}
{'='*60}

{site_settings.address if site_settings and site_settings.address else ''}

Employee Details:
- Name: {payslip.staff.user.get_full_name()}
- ID: {payslip.staff.employee_id}
- Position: {payslip.staff.position}

Earnings:
- Base Salary: GH₵{payslip.base_salary:,.2f}
- Total Allowances: GH₵{payslip.total_allowances:,.2f}
- Gross Pay: GH₵{payslip.gross_pay:,.2f}

Deductions:
- Total Deductions: GH₵{payslip.total_deductions:,.2f}

Net Pay: GH₵{payslip.net_pay:,.2f}

Payment Status: {payslip.get_status_display()}
Payment Date: {payslip.payment_date or 'Pending'}
Payment Reference: {payslip.payment_reference or 'Pending'}
"""
    
    response.write(content)
    return response
