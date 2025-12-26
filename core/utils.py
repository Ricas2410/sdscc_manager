"""
Utility functions for SDSCC system.
Includes: Export helpers, notification creation, report generation.
"""

import io
from datetime import date, timedelta
from decimal import Decimal
from django.http import HttpResponse
from django.utils import timezone
from django.db.models import Sum, Count, Avg


# ============ NOTIFICATION HELPERS ============

def create_notification(recipient, notification_type, title, message, link='', branch=None, created_by=None):
    """Create a notification for a user."""
    from .models import Notification
    
    return Notification.objects.create(
        recipient=recipient,
        notification_type=notification_type,
        title=title,
        message=message,
        link=link,
        branch=branch,
        created_by=created_by
    )


def notify_admins(branch, notification_type, title, message, link=''):
    """Send notification to all admins of a branch hierarchy."""
    from .models import Notification
    from accounts.models import User
    
    notifications = []
    
    # Branch executives
    branch_admins = User.objects.filter(
        branch=branch,
        role__in=['branch_executive', 'pastor']
    )
    
    # District executives
    if branch.district:
        district_admins = User.objects.filter(
            managed_district=branch.district,
            role='district_executive'
        )
        branch_admins = branch_admins | district_admins
    
    # Area executives
    if branch.district and branch.district.area:
        area_admins = User.objects.filter(
            managed_area=branch.district.area,
            role='area_executive'
        )
        branch_admins = branch_admins | area_admins
    
    # Mission admins
    mission_admins = User.objects.filter(role='mission_admin')
    all_admins = (branch_admins | mission_admins).distinct()
    
    for admin in all_admins:
        notifications.append(
            Notification(
                recipient=admin,
                notification_type=notification_type,
                title=title,
                message=message,
                link=link,
                branch=branch
            )
        )
    
    if notifications:
        Notification.objects.bulk_create(notifications)
    
    return len(notifications)


# ============ EXCEL EXPORT HELPERS ============

def export_to_excel(queryset, columns, filename, sheet_name='Data'):
    """
    Export a queryset to Excel file.
    
    Args:
        queryset: Django queryset or list of dicts
        columns: List of tuples (field_name, header_name)
        filename: Name of the file (without extension)
        sheet_name: Name of the Excel sheet
    
    Returns:
        HttpResponse with Excel file
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        return HttpResponse("openpyxl not installed", status=500)
    
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    
    # Header styles
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='1E40AF', end_color='1E40AF', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Write headers
    for col_idx, (field, header) in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Write data
    for row_idx, item in enumerate(queryset, 2):
        for col_idx, (field, header) in enumerate(columns, 1):
            if isinstance(item, dict):
                value = item.get(field, '')
            else:
                # Handle nested fields like 'branch.name'
                value = item
                for attr in field.split('.'):
                    if value is not None:
                        value = getattr(value, attr, None)
                        if callable(value):
                            value = value()
            
            # Format special types
            if isinstance(value, Decimal):
                value = float(value)
            elif isinstance(value, (date, timezone.datetime)):
                value = value.strftime('%Y-%m-%d')
            
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
    
    # Auto-adjust column widths
    for col_idx, (field, header) in enumerate(columns, 1):
        column_letter = get_column_letter(col_idx)
        max_length = len(header)
        for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[column_letter].width = min(max_length + 2, 50)
    
    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}.xlsx"'
    
    wb.save(response)
    return response


# ============ PDF EXPORT HELPERS ============

def export_to_pdf(title, headers, data, filename, orientation='portrait'):
    """
    Export data to PDF file.
    
    Args:
        title: Document title
        headers: List of column headers
        data: List of lists (rows of data)
        filename: Name of the file (without extension)
        orientation: 'portrait' or 'landscape'
    
    Returns:
        HttpResponse with PDF file
    """
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import letter, landscape
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import inch
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    except ImportError:
        return HttpResponse("reportlab not installed", status=500)
    
    buffer = io.BytesIO()
    
    pagesize = landscape(letter) if orientation == 'landscape' else letter
    doc = SimpleDocTemplate(buffer, pagesize=pagesize)
    
    elements = []
    styles = getSampleStyleSheet()
    
    # Title
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        spaceAfter=30,
        alignment=1  # Center
    )
    elements.append(Paragraph(title, title_style))
    
    # Subtitle with date
    subtitle_style = ParagraphStyle(
        'Subtitle',
        parent=styles['Normal'],
        fontSize=10,
        spaceAfter=20,
        alignment=1
    )
    elements.append(Paragraph(f"Generated on {date.today().strftime('%B %d, %Y')}", subtitle_style))
    
    # Table
    table_data = [headers] + data
    table = Table(table_data, repeatRows=1)
    
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1E40AF')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.white),
        ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#E5E7EB')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F9FAFB')]),
    ]))
    
    elements.append(table)
    
    doc.build(elements)
    
    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}.pdf"'
    
    return response


# ============ BIRTHDAY/ANNIVERSARY HELPERS ============

def get_upcoming_birthdays(branch=None, days=7):
    """Get members with birthdays in the next N days."""
    from accounts.models import User
    
    today = date.today()
    upcoming = []
    
    users = User.objects.filter(is_active=True)
    if branch:
        users = users.filter(branch=branch)
    
    for user in users.select_related('branch'):
        if user.date_of_birth:
            # Check if birthday falls within next N days
            bday_this_year = user.date_of_birth.replace(year=today.year)
            if bday_this_year < today:
                bday_this_year = bday_this_year.replace(year=today.year + 1)
            
            days_until = (bday_this_year - today).days
            if 0 <= days_until <= days:
                upcoming.append({
                    'user': user,
                    'date': bday_this_year,
                    'days_until': days_until,
                    'age': today.year - user.date_of_birth.year
                })
    
    return sorted(upcoming, key=lambda x: x['days_until'])


def get_upcoming_anniversaries(branch=None, days=7):
    """Get members with wedding anniversaries in the next N days."""
    from .models import SpecialDateReminder
    
    today = date.today()
    end_date = today + timedelta(days=days)
    upcoming = []
    
    reminders = SpecialDateReminder.objects.filter(
        reminder_type='wedding'
    ).select_related('user', 'user__branch')
    
    if branch:
        reminders = reminders.filter(user__branch=branch)
    
    for reminder in reminders:
        anniv_this_year = reminder.date.replace(year=today.year)
        if anniv_this_year < today:
            anniv_this_year = anniv_this_year.replace(year=today.year + 1)
        
        days_until = (anniv_this_year - today).days
        if 0 <= days_until <= days:
            years = today.year - (reminder.year or reminder.date.year)
            upcoming.append({
                'user': reminder.user,
                'date': anniv_this_year,
                'days_until': days_until,
                'years': years
            })
    
    return sorted(upcoming, key=lambda x: x['days_until'])


# ============ CSV IMPORT HELPERS ============

def parse_csv_members(csv_file, branch):
    """
    Parse a CSV file for member import.
    
    Expected columns:
    - first_name (required)
    - last_name (required)
    - phone (optional)
    - email (optional)
    - date_of_birth (optional, format: YYYY-MM-DD)
    - gender (optional, M/F)
    - address (optional)
    - occupation (optional)
    
    Returns:
        tuple: (success_list, error_list)
    """
    import csv
    from io import StringIO
    from accounts.models import User
    from django.conf import settings
    
    success = []
    errors = []
    
    # Read CSV
    try:
        content = csv_file.read().decode('utf-8')
        reader = csv.DictReader(StringIO(content))
    except Exception as e:
        return [], [{'row': 0, 'error': f'Failed to read CSV: {str(e)}'}]
    
    # Get SDSCC settings
    sdscc_settings = getattr(settings, 'SDSCC_SETTINGS', {})
    default_password = sdscc_settings.get('DEFAULT_PASSWORD', '12345')
    member_prefix = sdscc_settings.get('MEMBER_ID_PREFIX', 'MEM')
    
    for row_num, row in enumerate(reader, 2):  # Start at 2 (1 is header)
        try:
            first_name = row.get('first_name', '').strip()
            last_name = row.get('last_name', '').strip()
            
            if not first_name or not last_name:
                errors.append({'row': row_num, 'error': 'First name and last name are required'})
                continue
            
            # Generate member ID
            last_member = User.objects.filter(
                member_id__startswith=member_prefix
            ).order_by('-member_id').first()
            
            if last_member:
                try:
                    last_num = int(last_member.member_id.replace(member_prefix, ''))
                    new_num = last_num + 1
                except ValueError:
                    new_num = 1
            else:
                new_num = 1
            
            member_id = f"{member_prefix}{new_num:05d}"
            
            # Parse date of birth
            dob = None
            if row.get('date_of_birth'):
                try:
                    dob = date.fromisoformat(row['date_of_birth'].strip())
                except ValueError:
                    pass
            
            # Parse gender
            gender = row.get('gender', '').strip().upper()
            if gender not in ['M', 'F']:
                gender = ''
            
            # Create user
            user = User.objects.create_user(
                username=member_id,
                member_id=member_id,
                password=default_password,
                first_name=first_name,
                last_name=last_name,
                email=row.get('email', '').strip(),
                phone=row.get('phone', '').strip(),
                date_of_birth=dob,
                gender=gender,
                address=row.get('address', '').strip(),
                occupation=row.get('occupation', '').strip(),
                branch=branch,
                role='member',
                is_active=True
            )
            
            success.append({
                'row': row_num,
                'member_id': member_id,
                'name': f"{first_name} {last_name}"
            })
            
        except Exception as e:
            errors.append({'row': row_num, 'error': str(e)})
    
    return success, errors


# ============ MEMBER STATEMENT GENERATION ============

def generate_contribution_statement(user, year=None):
    """Generate a contribution statement PDF for a member."""
    from contributions.models import Contribution
    
    if year is None:
        year = date.today().year
    
    # Get contributions
    contributions = Contribution.objects.filter(
        member=user,
        date__year=year
    ).select_related('contribution_type').order_by('date')
    
    # Calculate totals by type
    totals_by_type = contributions.values(
        'contribution_type__name'
    ).annotate(
        total=Sum('amount'),
        count=Count('id')
    ).order_by('-total')
    
    grand_total = contributions.aggregate(total=Sum('amount'))['total'] or Decimal('0')
    
    # Prepare data
    headers = ['Date', 'Type', 'Amount (GH₵)', 'Reference']
    data = []
    for c in contributions:
        data.append([
            c.date.strftime('%Y-%m-%d'),
            c.contribution_type.name,
            f"{c.amount:,.2f}",
            c.reference or '-'
        ])
    
    # Add summary rows
    data.append(['', '', '', ''])
    data.append(['', 'SUMMARY BY TYPE', '', ''])
    for item in totals_by_type:
        data.append(['', item['contribution_type__name'], f"{item['total']:,.2f}", f"({item['count']} entries)"])
    
    data.append(['', '', '', ''])
    data.append(['', 'GRAND TOTAL', f"{grand_total:,.2f}", ''])
    
    title = f"Contribution Statement - {user.get_full_name()}\n{year}"
    
    return export_to_pdf(title, headers, data, f"statement_{user.member_id}_{year}")
